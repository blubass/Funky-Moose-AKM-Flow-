import json
import os
import tempfile
import unittest
from types import SimpleNamespace
from unittest import mock

from openpyxl import Workbook, load_workbook

import akm_app
from app_logic import akm_core
from app_logic import assistant_tools
from app_logic import cover_tools
from app_logic import detail_tools
from app_logic import flow_tools
from app_logic import loudness_tools
from app_logic.models import TrackRecord
from app_logic import overview_tools
from app_logic import release_tools
from app_logic import text_utils
from app_logic import akm_core as package_akm_core
from app_controllers import detail_controller_tools
from app_ui import cover_action_tools
from app_ui import cover_preview_tools
from app_ui import cover_view_tools
from app_ui import detail_view_tools
from app_ui import loudness_view_tools
from app_ui import path_ui_tools
from app_ui import release_view_tools
from app_ui import ui_patterns
from app_ui import widgets
from app_ui.tabs.batch_tab import BatchTab
from app_workflows import loudness_workflows
from app_workflows import release_workflows
from app_controllers.batch_controller import BatchController
from app_controllers.base_controller import BaseController
from app_controllers.details_controller import DetailsController
from app_controllers.loudness_controller import LoudnessController
from app_controllers.overview_controller import OverviewController
from app_controllers.project_controller import ProjectController
from app_controllers.release_controller import ReleaseController


class FakeVar:
    def __init__(self, value=""):
        self.value = value

    def get(self):
        return self.value

    def set(self, value):
        self.value = value


class BaseControllerTests(unittest.TestCase):
    def test_get_built_tab_prefers_app_hook(self):
        app = SimpleNamespace(
            state=SimpleNamespace(),
            tasks=SimpleNamespace(),
            get_built_tab=lambda tab_id: {"release": "hooked"}.get(tab_id),
            tab_system=SimpleNamespace(_instances={"release": "fallback"}),
        )

        controller = BaseController(app)

        self.assertEqual("hooked", controller.get_built_tab("release"))

    def test_get_built_tab_falls_back_to_tab_instances(self):
        app = SimpleNamespace(
            state=SimpleNamespace(),
            tasks=SimpleNamespace(),
            tab_system=SimpleNamespace(_instances={"release": "fallback"}),
        )

        controller = BaseController(app)

        self.assertEqual("fallback", controller.get_built_tab("release"))


class FakeText:
    def __init__(self, value=""):
        self.value = value

    def get(self, *_args):
        return self.value

    def delete(self, *_args):
        self.value = ""

    def insert(self, *_args):
        if _args:
            self.value = _args[-1]


class FakeEntry:
    def __init__(self):
        self.value = ""
        self.focused = False

    def delete(self, *_args):
        self.value = ""

    def insert(self, _index, value):
        self.value = value

    def get(self):
        return self.value

    def focus_set(self):
        self.focused = True


class FakeTabs:
    def __init__(self, tab_ids=None):
        self.selected = None
        self.tab_ids = list(tab_ids or [])
        self.updated = False
        self.focused = False

    def select(self, value):
        self.selected = value

    def tabs(self):
        return list(self.tab_ids)

    def update_idletasks(self):
        self.updated = True

    def focus_set(self):
        self.focused = True


class FakeProgress:
    def __init__(self):
        self.options = {}

    def __setitem__(self, key, value):
        self.options[key] = value

    def __getitem__(self, key):
        return self.options.get(key)


class FakeLabel:
    def __init__(self):
        self.options = {}

    def config(self, **kwargs):
        self.options.update(kwargs)


class FakePackWidget:
    def __init__(self):
        self.pack_calls = []
        self.pack_forget_calls = 0

    def pack(self, **kwargs):
        self.pack_calls.append(kwargs)

    def pack_forget(self):
        self.pack_forget_calls += 1


class FakeBatchView:
    def __init__(self, copy_stage="title"):
        self.flow_state = None
        self.empty_rendered = False
        self.copy_button_text = None
        self.copy_stage = copy_stage

    def render_flow_state(self, **kwargs):
        self.flow_state = kwargs

    def render_empty_state(self):
        self.empty_rendered = True

    def set_copy_button_label(self, text):
        self.copy_button_text = text

    def get_copy_stage(self):
        return self.copy_stage

    def set_copy_stage(self, stage):
        self.copy_stage = stage


class FakeConfigButton:
    def __init__(self):
        self.options = {}

    def config(self, **kwargs):
        self.options.update(kwargs)


class FakeMessagebox:
    def __init__(self, response):
        self.response = response

    def askyesnocancel(self, *_args, **_kwargs):
        return self.response


class FakeReleaseListbox:
    def __init__(self, selection=()):
        self.selection = tuple(selection)
        self.selected_index = None
        self.items = []

    def curselection(self):
        return self.selection

    def selection_set(self, index):
        self.selected_index = index

    def delete(self, *_args):
        self.items = []

    def insert(self, _index, value):
        self.items.append(value)


class FakeReleaseView:
    def __init__(self, selection=(), form_state=None):
        self.selection = tuple(selection)
        self.selected_index = None
        self.track_labels = []
        self.action_hint = None
        self.preflight_text = None
        self.flow_hint = None
        self.status_text = None
        self.form_state = {
            "title": "",
            "artist": "",
            "type": "",
            "cover_path": "",
            "export_dir": "",
        }
        if form_state:
            self.form_state.update(form_state)

    def has_track_list(self):
        return True

    def render_release_state(self, track_labels, action_hint, preflight_text, flow_hint, status_text):
        self.track_labels = list(track_labels)
        self.action_hint = action_hint
        self.preflight_text = preflight_text
        self.flow_hint = flow_hint
        self.status_text = status_text

    def get_selected_track_indices(self):
        return self.selection

    def select_track_index(self, index):
        self.selected_index = index

    def get_form_state(self):
        return dict(self.form_state)

    def set_form_state(self, values):
        self.form_state.update(values or {})

    def get_form_value(self, key):
        return (self.form_state.get(key) or "").strip()

    def set_form_value(self, key, value):
        self.form_state[key] = value


class FakeOverviewView:
    def __init__(self, selection=(), filter_state=None):
        self.selection = tuple(selection)
        self.list_items = []
        self.row_statuses = []
        self.summary_text = None
        self.status_text = None
        self.hint_text = None
        self.empty_text = None
        self.empty_visible = None
        self.filter_chip_texts = {}
        self.filter_chip_selected = {}
        self.filter_state = {
            "search": "",
            "filter": "all",
            "sort": "title",
            "desc": False,
        }
        if filter_state:
            self.filter_state.update(filter_state)

    def render_overview_records(self, labels, row_statuses):
        self.list_items = list(labels)
        self.row_statuses = list(row_statuses)

    def render_overview_meta(self, *, summary_text, status_text, hint_text, empty_text, show_empty):
        self.summary_text = summary_text
        self.status_text = status_text
        self.hint_text = hint_text
        self.empty_text = empty_text
        self.empty_visible = show_empty

    def update_filter_chip(self, status_key, text, selected=False):
        self.filter_chip_texts[status_key] = text
        self.filter_chip_selected[status_key] = selected

    def get_filter_state(self):
        return dict(self.filter_state)

    def set_status_filter(self, status_key):
        self.filter_state["filter"] = status_key or "all"

    def get_selected_indices(self):
        return self.selection


class FakeDashboardView:
    def __init__(self):
        self.stats = None
        self.status_text = None
        self.hint_text = None
        self.meta_text = None
        self.chip_counts = None
        self.chip_labels = {}

    def render_dashboard_state(self, stats, status_text, hint_text, meta_text, chip_counts, status_text_fn):
        self.stats = dict(stats)
        self.status_text = status_text
        self.hint_text = hint_text
        self.meta_text = meta_text
        self.chip_counts = dict(chip_counts)
        self.chip_labels = {
            key: f"{status_text_fn(key)}  {chip_counts.get(key, 0)}"
            for key in chip_counts
        }


class FakeDetailsView:
    def __init__(self, form_vars=None):
        self.title_values = []
        self.form_vars = dict(form_vars or {})
        self.notes_text = ""
        self.tags_text = ""
        self.instrumental = False
        self.status_key = None
        self.status_text = None
        self.refresh_calls = 0

    def refresh_view(self):
        self.refresh_calls += 1

    def get_form_vars(self):
        return self.form_vars

    def set_title_options(self, titles):
        self.title_values = list(titles)

    def get_notes_text(self):
        return self.notes_text

    def set_notes_text(self, text):
        self.notes_text = text or ""

    def clear_notes(self):
        self.notes_text = ""

    def get_tags_text(self):
        return self.tags_text

    def set_tags_text(self, text):
        self.tags_text = text or ""

    def clear_tags(self):
        self.tags_text = ""

    def get_instrumental(self):
        return self.instrumental

    def set_instrumental(self, value):
        self.instrumental = bool(value)

    def set_status_chip_display(self, status_key, status_label):
        self.status_key = status_key
        self.status_text = status_label


class FakeLoudnessView:
    def __init__(
        self,
        *,
        target_text="-14.0",
        peak_text="-1.0",
        output_dir="",
        use_limiter=True,
        selection=(),
    ):
        self.target_text = target_text
        self.peak_text = peak_text
        self.output_dir = output_dir
        self.use_limiter = use_limiter
        self.selection = tuple(selection)
        self.status_text = None
        self.hint_text = None
        self.log_text = ""
        self.rows = []

    def has_tree(self):
        return True

    def get_target_text(self):
        return self.target_text

    def get_peak_text(self):
        return self.peak_text

    def get_output_dir(self):
        return self.output_dir

    def set_output_dir(self, path):
        self.output_dir = path or ""

    def get_use_limiter(self):
        return self.use_limiter

    def get_selected_paths(self):
        return self.selection

    def clear_tree(self):
        self.rows = []

    def insert_tree_row(self, path, values, tags=()):
        self.rows.append({"path": path, "values": tuple(values), "tags": tuple(tags)})

    def apply_workflow_state(self, status_text="", hint_text="", log_lines=None):
        self.status_text = status_text or ""
        self.hint_text = hint_text or ""
        self.log_text = "\n".join(log_lines or [])

    def set_open_state(self, status_text="", hint_text=""):
        self.status_text = status_text or ""
        self.hint_text = hint_text or ""


class FakeListbox:
    def __init__(self, selection=()):
        self.selection = tuple(selection)
        self.items = []
        self.row_options = {}

    def delete(self, *_args):
        self.items = []

    def insert(self, _index, *values):
        self.items.extend(values)

    def itemconfig(self, index, **kwargs):
        self.row_options[index] = kwargs

    def curselection(self):
        return self.selection


class FakeWidgetNode:
    def __init__(self, name, registry, parent=None):
        self._name = name
        self._registry = registry
        self._parent = parent
        registry[name] = self

    def winfo_parent(self):
        return "" if self._parent is None else self._parent._name

    def nametowidget(self, name):
        return self._registry[name]


class FakeNativeScrollWidget(FakeWidgetNode):
    def __init__(self, name, registry, parent=None):
        super().__init__(name, registry, parent=parent)
        self.scroll_calls = []

    def yview_scroll(self, amount, units):
        self.scroll_calls.append((amount, units))


class FakeCanvas:
    def __init__(self):
        self.scroll_calls = []

    def yview_scroll(self, amount, units):
        self.scroll_calls.append((amount, units))


class FakeScrollablePanelWidget(FakeWidgetNode):
    def __init__(self, name, registry, parent=None):
        super().__init__(name, registry, parent=parent)
        self.canvas = FakeCanvas()


class ImmediateTaskRunner:
    def run(self, task_func, on_success=None, on_error=None, busy_text=None):
        del busy_text
        try:
            result = task_func()
        except Exception as exc:
            if on_error:
                on_error(str(exc))
                return
            raise
        if on_success:
            on_success(result)


class FakeState:
    def __init__(self, records=None, mtime=None):
        self._records = list(records or [])
        self.filtered_records = []
        self.batch_queue = []
        self.batch_index = 0
        self.release_tracks = []
        self.loudness_files = []
        self.loudness_results = []
        self.current_mtime = mtime
        self.invalidated = False

    def get_all_records(self, force=False, copy_data=False):
        del force
        if copy_data:
            return [dict(item) for item in self._records]
        return self._records

    def invalidate_cache(self):
        self.invalidated = True

    def _get_data_mtime(self):
        return self.current_mtime


class FakeLoudnessModule:
    @staticmethod
    def calculate_gain_to_target(measured_lufs, target_lufs):
        if measured_lufs is None:
            return None
        return round(target_lufs - measured_lufs, 2)

    @staticmethod
    def predict_true_peak_after_gain(current_true_peak_dbtp, gain_db):
        if current_true_peak_dbtp is None or gain_db is None:
            return None
        return round(current_true_peak_dbtp + gain_db, 2)

    @staticmethod
    def get_status_for_match(measured_lufs, gain_db, predicted_true_peak, true_peak_ceiling=-1.0):
        if measured_lufs is None or gain_db is None:
            return "Analyse fehlt"
        if predicted_true_peak is not None and predicted_true_peak > true_peak_ceiling:
            return "Peak Warnung"
        if abs(gain_db) < 0.3:
            return "OK"
        return "Zu leise" if gain_db > 0 else "Zu laut"

    @staticmethod
    def safe_output_path(output_dir, source_path, suffix="_matched"):
        stem, ext = os.path.splitext(os.path.basename(source_path))
        return os.path.join(output_dir, f"{stem}{suffix}{ext}")

    @staticmethod
    def export_matched_file(source_path, output_path, gain_db, use_limiter=False, true_peak_ceiling_db=-1.0):
        return {
            "ok": True,
            "source": source_path,
            "output": output_path,
            "gain_db": gain_db,
            "use_limiter": use_limiter,
            "true_peak_ceiling_db": true_peak_ceiling_db,
            "error": "",
        }


class TemporaryStorageTestCase(unittest.TestCase):
    def setUp(self):
        super().setUp()
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        patch_values = dict(
            DATA_DIR=self.tempdir.name,
            DATA_FILE=os.path.join(self.tempdir.name, "data.json"),
            BACKUP_FILE=os.path.join(self.tempdir.name, "data_backup.json"),
            LANG_FILE=os.path.join(self.tempdir.name, "lang.txt"),
            SETTINGS_FILE=os.path.join(self.tempdir.name, "settings.json"),
        )
        self.storage_patcher = mock.patch.multiple(
            akm_core,
            **patch_values,
        )
        self.storage_patcher.start()
        self.addCleanup(self.storage_patcher.stop)
        self.package_storage_patcher = mock.patch.multiple(
            package_akm_core,
            **patch_values,
        )
        self.package_storage_patcher.start()
        self.addCleanup(self.package_storage_patcher.stop)

    def save_entries(self, entries):
        akm_core.save_data(entries)

    def load_entries(self):
        return akm_core.load_data(strict=True)


class ExcelImportTests(TemporaryStorageTestCase):
    def test_load_excel_tracks_rejects_unsupported_extension(self):
        invalid_path = os.path.join(self.tempdir.name, "tracks.xls")
        with open(invalid_path, "w", encoding="utf-8") as handle:
            handle.write("not a workbook")

        with self.assertRaisesRegex(
            akm_core.DataFileError,
            "Nicht unterstütztes Excel-Format",
        ):
            akm_core.load_excel_tracks(invalid_path)

    def test_load_excel_tracks_reads_supported_workbook(self):
        workbook_path = os.path.join(self.tempdir.name, "tracks.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Titel", "Dauer", "Komponist"])
        sheet.append(["Mein Song", "3:21", "Uwe"])
        workbook.save(workbook_path)
        workbook.close()

        tracks = akm_core.load_excel_tracks(workbook_path)

        self.assertEqual(1, len(tracks))
        self.assertEqual("Mein Song", tracks[0]["title"])
        self.assertEqual("3:21", tracks[0]["duration"])
        self.assertEqual("Uwe", tracks[0]["composer"])

    def test_load_excel_tracks_reads_release_export_import_sheet(self):
        workbook_path = os.path.join(self.tempdir.name, "release.xlsx")
        release_workflows.write_release_track_workbook(
            workbook_path,
            {
                "title": "My Release",
                "artist": "Artist",
                "type": "Album",
                "release_date": "2026-04-09",
                "genre": "Pop",
                "subgenre": "Indie",
                "label": "Label",
                "copyright_line": "C 2026",
                "cover_path": "/tmp/cover.jpg",
            },
            [
                {
                    "title": "Mein Song",
                    "duration": "3:21",
                    "composer": "Uwe",
                    "production": "FM",
                    "year": "2026",
                    "notes": "Single Mix",
                    "audio_path": "/tmp/song.wav",
                    "source": "Werk",
                }
            ],
        )

        tracks = akm_core.load_excel_tracks(workbook_path)

        self.assertEqual(1, len(tracks))
        self.assertEqual("Mein Song", tracks[0]["title"])
        self.assertEqual("3:21", tracks[0]["duration"])
        self.assertEqual("Uwe", tracks[0]["composer"])
        self.assertEqual("/tmp/song.wav", tracks[0]["audio_path"])


class SettingsMemoryTests(TemporaryStorageTestCase):
    def test_settings_helpers_persist_detail_and_release_memory(self):
        akm_core.remember_detail_memory(
            {
                "composer": "Uwe",
                "production": "FM",
                "year": "2026",
            }
        )
        akm_core.remember_release_memory(
            {
                "artist": "Artist X",
                "type": "Album",
                "genre": "Pop",
                "subgenre": "Indie",
                "label": "Label X",
                "copyright_line": "C 2026",
                "export_dir": "/tmp/exports",
            }
        )

        self.assertEqual(
            {
                "composer": "Uwe",
                "production": "FM",
                "year": "2026",
            },
            akm_core.get_detail_memory(),
        )
        self.assertEqual("Artist X", akm_core.get_release_memory()["artist"])
        self.assertEqual("Album", akm_core.get_release_memory()["type"])
        self.assertEqual("/tmp/exports", akm_core.get_release_memory()["export_dir"])
        self.assertEqual("Artist X", akm_core.get_release_default_artist())


class OverviewToolsTests(unittest.TestCase):
    def test_build_dashboard_stats_counts_key_fields(self):
        entries = [
            {"status": "in_progress", "instrumental": True, "production": "P1", "notes": ""},
            {"status": "ready", "instrumental": False, "production": "", "notes": "note"},
            {"status": "confirmed", "instrumental": False, "production": "", "notes": ""},
        ]

        stats = overview_tools.build_dashboard_stats(entries)

        self.assertEqual(3, stats["total"])
        self.assertEqual(2, stats["open"])
        self.assertEqual(1, stats["in_progress"])
        self.assertEqual(1, stats["ready"])
        self.assertEqual(1, stats["confirmed"])
        self.assertEqual(1, stats["instrumental"])
        self.assertEqual(1, stats["with_production"])
        self.assertEqual(1, stats["with_notes"])

    def test_dashboard_helpers_build_status_focus_and_meta_texts(self):
        stats = {
            "total": 8,
            "open": 5,
            "confirmed": 3,
            "ready": 2,
            "submitted": 1,
            "in_progress": 2,
            "with_production": 5,
            "with_notes": 4,
            "instrumental": 1,
        }

        self.assertEqual(38, overview_tools.build_dashboard_completion_percent(stats))
        self.assertEqual(
            "8 Werke im Katalog   •   38% bestätigt   •   2 bereit   •   2 in Arbeit",
            overview_tools.build_dashboard_status_text(stats),
        )
        self.assertIn("2 Werk(e) brauchen noch Feinschliff", overview_tools.build_dashboard_focus_text(stats))
        self.assertEqual(
            "Offen: 5   •   Mit Produktion: 5   •   Mit Notizen: 4   •   Instrumental: 1",
            overview_tools.build_dashboard_meta_text(stats),
        )

    def test_filter_and_sort_entries_applies_query_filter_and_sorting(self):
        entries = [
            {"title": "Beta", "status": "confirmed", "year": "2021", "notes": "", "tags": []},
            {"title": "Alpha", "status": "ready", "year": "2024", "notes": "Synthwave", "tags": ["focus"]},
            {"title": "Gamma", "status": "in_progress", "year": "2023", "notes": "", "tags": []},
        ]

        filtered = overview_tools.filter_and_sort_entries(
            entries,
            query="focus",
            status_filter="open",
            sort_key="year",
            sort_desc=True,
        )

        self.assertEqual(["Alpha"], [item["title"] for item in filtered])

    def test_build_overview_summary_formats_human_readable_text(self):
        summary = overview_tools.build_overview_summary(
            5,
            status_filter="ready",
            query="mix",
            sort_key="last_change",
            sort_desc=True,
            status_label="Bereit",
        )

        self.assertEqual(
            "5 Treffer   •   Status: Bereit   •   Suche: mix   •   Sortierung: Änderung absteigend",
            summary,
        )

    def test_overview_helpers_describe_empty_and_filtered_states(self):
        self.assertEqual(
            "Noch keine Werke im Katalog",
            overview_tools.build_overview_status_text(0, 0),
        )
        self.assertEqual(
            "Keine Treffer im aktuellen Filter",
            overview_tools.build_overview_status_text(0, 8),
        )
        self.assertIn(
            "keine Treffer",
            overview_tools.build_overview_hint_text(0, 4, status_filter="ready", query="mix"),
        )


class DetailToolsTests(unittest.TestCase):
    def test_detail_form_state_from_item_normalizes_for_widgets(self):
        state = detail_tools.detail_form_state_from_item(
            {
                "title": "Song",
                "duration": "3:21",
                "composer": "Uwe",
                "production": "FM",
                "year": "2026",
                "audio_path": "/tmp/song.wav",
                "instrumental": 1,
                "status": "ready",
                "notes": "Notiz",
                "tags": [" synth ", float("nan"), "", "focus"],
            }
        )

        self.assertEqual("Song", state["original_title"])
        self.assertEqual("Song", state["current_title"])
        self.assertEqual("3:21", state["values"]["duration"])
        self.assertTrue(state["instrumental"])
        self.assertEqual("ready", state["status"])
        self.assertEqual("synth, focus", state["tags_text"])
        self.assertEqual("Notiz", state["notes"])

    def test_build_detail_updates_trims_text_and_splits_tags(self):
        updates = detail_tools.build_detail_updates(
            {
                "title": "  Song  ",
                "duration": " 3:33 ",
                "composer": " Uwe ",
                "production": " Prod ",
                "year": " 2026 ",
                "audio_path": " /tmp/song.wav ",
            },
            " synth, , focus ,",
            "  Notiz  ",
            "submitted",
            1,
        )

        self.assertEqual("Song", updates["title"])
        self.assertEqual("3:33", updates["duration"])
        self.assertEqual("Uwe", updates["composer"])
        self.assertEqual("Prod", updates["production"])
        self.assertEqual("2026", updates["year"])
        self.assertEqual("/tmp/song.wav", updates["audio_path"])
        self.assertEqual("submitted", updates["status"])
        self.assertTrue(updates["instrumental"])
        self.assertEqual("Notiz", updates["notes"])
        self.assertEqual(["synth", "focus"], updates["tags"])

    def test_resolve_original_title_prefers_loaded_then_existing_then_current(self):
        self.assertEqual(
            "Loaded Song",
            detail_tools.resolve_original_title(
                "Loaded Song",
                "Renamed Song",
                [{"title": "Renamed Song"}],
            ),
        )
        self.assertEqual(
            "Existing Song",
            detail_tools.resolve_original_title(
                "",
                "Existing Song",
                [{"title": "Existing Song"}],
            ),
        )
        self.assertEqual(
            "Fresh Song",
            detail_tools.resolve_original_title(
                None,
                "Fresh Song",
                [{"title": "Other Song"}],
            ),
        )


class AssistantToolsTests(unittest.TestCase):
    def test_build_import_log_messages_formats_summary_and_rows(self):
        messages = assistant_tools.build_import_log_messages(
            [
                {
                    "title": "Song A",
                    "action": "added",
                    "duration": "3:11",
                    "composer": "Uwe",
                },
                {
                    "title": "Song B",
                    "action": "unchanged",
                    "duration": float("nan"),
                    "composer": " nan ",
                },
            ]
        )

        self.assertEqual("Excel-Import abgeschlossen: 2 Einträge", messages[0])
        self.assertEqual("  + Song A (3:11 | Uwe)", messages[1])
        self.assertEqual("  = Song B", messages[2])

    def test_build_status_and_add_messages_cover_success_and_errors(self):
        self.assertEqual(
            "Neu angelegt: Song A",
            assistant_tools.build_add_result_message("Song A", True, {"title": "Song A"}),
        )
        self.assertEqual(
            "Schon vorhanden: Song A",
            assistant_tools.build_add_result_message("Song A", False, "already_exists"),
        )
        self.assertEqual(
            "Status gesetzt: Song A -> Bereit",
            assistant_tools.build_status_result_message(
                "Song A",
                True,
                "Bereit",
                {"status": "ready"},
                "ready",
            ),
        )
        self.assertEqual(
            "Nicht gefunden: Song A",
            assistant_tools.build_status_result_message(
                "Song A",
                False,
                "Bereit",
                "not_found",
                "ready",
            ),
        )

    def test_build_loudness_tab_open_state_switches_by_tool_availability(self):
        available = assistant_tools.build_loudness_tab_open_state(True)
        missing = assistant_tools.build_loudness_tab_open_state(False)

        self.assertIn("Dateien wählen", available["status_text"])
        self.assertIn("konnte nicht geladen", missing["status_text"])
        self.assertEqual(
            "Dateien laden oder direkt Werke aus der aktuellen Übersicht übernehmen.",
            available["hint_text"],
        )

    def test_build_assistant_radar_state_handles_empty_and_filled_input(self):
        empty = assistant_tools.build_assistant_radar_state("")
        filled = assistant_tools.build_assistant_radar_state("Neuer Song")

        self.assertEqual("Schnellstart bereit", empty["status_text"])
        self.assertIn("Excel-Import", empty["meta_text"])
        self.assertEqual("Bereit für neuen Titel: Neuer Song", filled["status_text"])
        self.assertIn("2 Wort/Wörter", filled["meta_text"])


class TextUtilsTests(unittest.TestCase):
    def test_clean_text_and_mapping_values_handle_nan_like_input(self):
        values = text_utils.clean_mapping_values(
            {
                "title": "  Song A  ",
                "duration": float("nan"),
                "composer": " nan ",
            },
            ("title", "duration", "composer", "year"),
        )

        self.assertEqual("Song A", text_utils.clean_text("  Song A  "))
        self.assertEqual("", text_utils.clean_text(float("nan")))
        self.assertEqual("", text_utils.clean_text(" nan "))
        self.assertEqual(
            {
                "title": "Song A",
                "duration": "",
                "composer": "",
                "year": "",
            },
            values,
        )


class FlowToolsTests(unittest.TestCase):
    def test_resolve_flow_index_prefers_previous_title_then_clamps(self):
        entries = [
            {"title": "Alpha"},
            {"title": "Beta"},
            {"title": "Gamma"},
        ]

        self.assertEqual(
            1,
            flow_tools.resolve_flow_index(
                entries,
                current_index=0,
                previous_title="Beta",
                preferred_index=2,
            ),
        )
        self.assertEqual(
            2,
            flow_tools.resolve_flow_index(
                entries,
                current_index=1,
                previous_title="Missing",
                preferred_index=99,
            ),
        )

    def test_build_flow_state_formats_meta_progress_and_copy_label(self):
        state = flow_tools.build_flow_state(
            [
                {
                    "title": "Song A",
                    "composer": "Uwe",
                    "duration": "3:11",
                    "production": "FM",
                    "year": "2026",
                }
            ],
            0,
            "duration",
        )

        self.assertTrue(state["has_item"])
        self.assertEqual("Song A", state["title_text"])
        self.assertEqual(
            "Komponist: Uwe   Dauer: 3:11   Produktion: FM   Jahr: 2026",
            state["meta_text"],
        )
        self.assertEqual("1 / 1", state["progress_text"])
        self.assertEqual(100.0, state["progress_value"])
        self.assertEqual("Dauer kopieren", state["copy_button_label"])

    def test_resolve_copy_action_cycles_title_then_duration(self):
        item = {"title": "Song A", "duration": "3:11"}

        title_step = flow_tools.resolve_copy_action(item, "title")
        duration_step = flow_tools.resolve_copy_action(item, "duration")

        self.assertEqual("Song A", title_step["value"])
        self.assertEqual("duration", title_step["next_stage"])
        self.assertFalse(title_step["advance_after_copy"])
        self.assertEqual("3:11", duration_step["value"])
        self.assertEqual("title", duration_step["next_stage"])
        self.assertTrue(duration_step["advance_after_copy"])

    def test_next_flow_index_wraps_at_end(self):
        self.assertEqual(0, flow_tools.next_flow_index(2, 3))
        self.assertEqual(1, flow_tools.next_flow_index(0, 3))

    def test_flow_dashboard_helpers_describe_queue_and_next_action(self):
        entries = [
            {"title": "Song A", "status": "in_progress", "duration": "3:11"},
            {"title": "Song B", "status": "ready"},
        ]
        flow_state = flow_tools.build_flow_state(entries, 0, "title")

        self.assertEqual(
            {"in_progress": 1, "ready": 1},
            flow_tools.build_flow_queue_counts(entries),
        )
        self.assertEqual(
            "Fokus 1 / 2   •   Song A",
            flow_tools.build_flow_status_text(entries, flow_state),
        )
        self.assertIn(
            "Copy-Flow automatisch auf die Dauer",
            flow_tools.build_flow_hint_text(entries, flow_state, "title"),
        )
        self.assertEqual(
            "In Arbeit: 1   •   Bereit: 1   •   Offen im Flow: 2",
            flow_tools.build_flow_meta_summary(entries),
        )


class ReleaseWorkflowTests(unittest.TestCase):
    def test_clean_release_drop_paths_dedupes_and_strips_braces(self):
        paths = release_workflows.clean_release_drop_paths(
            ["{/tmp/a.wav}", " /tmp/b.wav ", "/tmp/a.wav", ""]
        )

        self.assertEqual(["/tmp/a.wav", "/tmp/b.wav"], paths)

    def test_append_unique_release_tracks_skips_duplicates(self):
        result = release_workflows.append_unique_release_tracks(
            [{"title": "Song A", "source": "Werk", "audio_path": "/a.wav"}],
            [
                {"title": "Song A", "source": "Werk", "audio_path": "/a-other.wav"},
                {"title": "Song B", "source": "Datei", "audio_path": "/b.wav"},
            ],
        )

        self.assertEqual(2, len(result["tracks"]))
        self.assertEqual("Song B", result["added"][0]["title"])
        self.assertEqual("Song A", result["duplicates"][0]["title"])

    def test_move_release_track_and_target_name_helpers(self):
        moved, new_index = release_workflows.move_release_track(
            [{"title": "A"}, {"title": "B"}],
            0,
            1,
        )
        target_name = release_workflows.build_release_audio_target_name(
            2,
            {"title": "My:Song?"},
            "/tmp/source.wav",
        )

        self.assertEqual(1, new_index)
        self.assertEqual("B", moved[0]["title"])
        self.assertEqual("02 - My_Song_.wav", target_name)

    def test_build_release_export_text_helpers(self):
        info_lines = release_workflows.build_release_info_lines(
            {
                "title": "Release",
                "artist": "Artist",
                "type": "Album",
                "release_date": "2026-04-05",
                "genre": "Pop",
                "subgenre": "Indie",
                "label": "Label",
                "copyright_line": "C 2026",
                "cover_path": "/tmp/cover.jpg",
            }
        )
        checklist = release_workflows.build_release_checklist_lines(
            "/tmp/cover.jpg",
            [{"title": "Song"}],
            1,
            0,
        )

        self.assertEqual("Release Title: Release", info_lines[0])
        self.assertEqual("Cover gesetzt: Ja", checklist[0])
        self.assertEqual("Tracks im Release: 1", checklist[1])


class PathUiToolsTests(unittest.TestCase):
    def test_validate_existing_path_and_message_helpers(self):
        normalized, error = path_ui_tools.validate_existing_path(
            "",
            "Kein Pfad gesetzt.",
            "Datei nicht gefunden",
        )
        found_path, found_error = path_ui_tools.validate_existing_path(
            "/tmp/missing.wav",
            "Kein Pfad gesetzt.",
            "Datei nicht gefunden",
            exists_fn=lambda _path: False,
        )

        self.assertEqual("", normalized)
        self.assertEqual("Kein Pfad gesetzt.", error)
        self.assertEqual("/tmp/missing.wav", found_path)
        self.assertEqual("Datei nicht gefunden: /tmp/missing.wav", found_error)
        self.assertEqual(
            "Audio-Pfad gesetzt: song.wav",
            path_ui_tools.build_file_set_message("Audio-Pfad gesetzt", "/tmp/song.wav"),
        )


class UiPatternsTests(unittest.TestCase):
    def test_apply_button_bar_layout_honors_custom_anchor_in_stack_and_row(self):
        container = FakePackWidget()
        buttons = (FakePackWidget(), FakePackWidget())

        mode = ui_patterns.apply_button_bar_layout(
            container,
            buttons,
            width=500,
            breakpoint=600,
            mode=None,
            row_spacing=7,
            anchor="e",
        )

        self.assertEqual("stack", mode)
        self.assertEqual({"anchor": "e", "fill": "x"}, container.pack_calls[-1])
        self.assertEqual({"fill": "x", "pady": (0, ui_patterns.SPACE_XS)}, buttons[0].pack_calls[-1])
        self.assertEqual({"fill": "x", "pady": (0, 0)}, buttons[1].pack_calls[-1])

        mode = ui_patterns.apply_button_bar_layout(
            container,
            buttons,
            width=700,
            breakpoint=600,
            mode=mode,
            row_spacing=7,
            anchor="e",
        )

        self.assertEqual("row", mode)
        self.assertEqual({"anchor": "e"}, container.pack_calls[-1])
        self.assertEqual({"side": "left", "padx": (7, 0)}, buttons[1].pack_calls[-1])

    def test_apply_widget_layout_repacks_each_widget_once_and_returns_mode(self):
        left = FakePackWidget()
        right = FakePackWidget()

        mode = ui_patterns.apply_widget_layout(
            width=540,
            breakpoint=700,
            mode=None,
            layout_map={
                "stack": (
                    (left, {"fill": "x", "pady": (0, 8)}),
                    (right, {"fill": "x"}),
                ),
                "row": (
                    (left, {"side": "left"}),
                    (right, {"side": "left"}),
                ),
            },
        )

        self.assertEqual("stack", mode)
        self.assertEqual(1, left.pack_forget_calls)
        self.assertEqual(1, right.pack_forget_calls)
        self.assertEqual({"fill": "x", "pady": (0, 8)}, left.pack_calls[-1])

        mode = ui_patterns.apply_widget_layout(
            width=540,
            breakpoint=700,
            mode=mode,
            layout_map={
                "stack": (
                    (left, {"fill": "x", "pady": (0, 8)}),
                    (right, {"fill": "x"}),
                ),
                "row": (
                    (left, {"side": "left"}),
                    (right, {"side": "left"}),
                ),
            },
        )

        self.assertEqual("stack", mode)
        self.assertEqual(1, left.pack_forget_calls)


class CoverActionToolsTests(unittest.TestCase):
    def test_analyze_cover_drop_files_prefers_first_supported_and_reports_ignored(self):
        selected = cover_action_tools.analyze_cover_drop_files(
            ["", "/tmp/cover.png", "/tmp/other.jpg"],
            find_supported_artwork_path_fn=lambda paths: next(
                (path for path in paths if path.endswith(".png")),
                "",
            ),
        )
        ignored = cover_action_tools.analyze_cover_drop_files(
            ["/tmp/readme.txt", "/tmp/track.wav"],
            find_supported_artwork_path_fn=lambda _paths: "",
        )

        self.assertEqual("/tmp/cover.png", selected["target_path"])
        self.assertEqual("", selected["ignored_message"])
        self.assertEqual("", ignored["target_path"])
        self.assertIn("2 Datei(en)", ignored["ignored_message"])

    def test_prepare_cover_load_assignment_and_export_requests_validate_paths(self):
        load_state = cover_action_tools.prepare_cover_load(
            " /tmp/cover.png ",
            is_supported_path_fn=lambda path: path == "/tmp/cover.png",
        )
        assignment = cover_action_tools.prepare_cover_assignment(
            "/tmp/cover.png",
            exists_fn=lambda path: path == "/tmp/cover.png",
        )
        export_request = cover_action_tools.build_cover_export_request(
            "/tmp/cover.png",
            "Live Cover",
            exists_fn=lambda path: path == "/tmp/cover.png",
        )

        self.assertTrue(load_state["ok"])
        self.assertEqual("/tmp/cover.png", load_state["path"])
        self.assertEqual("Cover geladen: cover.png", load_state["log_message"])
        self.assertTrue(assignment["ok"])
        self.assertEqual("/tmp/cover.png", assignment["path"])
        self.assertTrue(export_request["ok"])
        self.assertEqual("/tmp", export_request["dialog_options"]["initialdir"])
        self.assertEqual("Cover_Live_Cover.jpg", export_request["dialog_options"]["initialfile"])

    def test_build_cover_export_success_state_uses_exported_basename(self):
        state = cover_action_tools.build_cover_export_success_state(
            "/tmp/final-cover.png"
        )

        self.assertEqual("Cover exportiert: final-cover.png", state["log_message"])
        self.assertEqual("COVER ERFOLGREICH EXPORTIERT", state["toast_message"])


class CoverViewToolsTests(unittest.TestCase):
    def test_first_supported_artwork_path_picks_first_valid_image(self):
        with tempfile.TemporaryDirectory() as tempdir:
            invalid_path = os.path.join(tempdir, "notes.txt")
            valid_path = os.path.join(tempdir, "cover.png")
            with open(invalid_path, "w", encoding="utf-8") as handle:
                handle.write("x")
            with open(valid_path, "wb") as handle:
                handle.write(b"png")

            selected = cover_view_tools.first_supported_artwork_path(
                [invalid_path, valid_path]
            )

            self.assertEqual(valid_path, selected)

    def test_build_cover_dashboard_state_reports_loaded_render(self):
        with tempfile.TemporaryDirectory() as tempdir:
            artwork_path = os.path.join(tempdir, "cover.png")
            with open(artwork_path, "wb") as handle:
                handle.write(b"png")

            state = cover_view_tools.build_cover_dashboard_state(
                layout_value="top-left",
                style_value="bold",
                size_mode_value="large",
                overlay_value="strong",
                offset_value="low",
                preview_height=320,
                artwork_path=artwork_path,
                artwork_meta={
                    "dimensions": (3000, 3000),
                    "size_text": "1.2 MB",
                    "ext": "PNG",
                },
                current_image_size=(1800, 1800),
                preview_dimensions=(640, 360),
                is_rendering=False,
                last_preview_error="",
            )

            self.assertEqual("Top Left Card bereit fuer Export", state["status_text"])
            self.assertIn("Layout: Top Left Card | Stil: bold", state["meta_text"])
            self.assertIn("Render: 1800x1800", state["info_text"])
            self.assertIn("Stage: 640x360", state["info_text"])
            self.assertEqual("cover.png", state["asset_name"])
            self.assertIn("3000x3000 px", state["asset_meta"])


class CoverPreviewToolsTests(unittest.TestCase):
    def test_build_cover_render_payload_formats_text_and_manual_layers(self):
        payload = cover_preview_tools.build_cover_render_payload(
            {
                "artwork_path": " /tmp/cover.png ",
                "zoom": "1.25",
                "layout": "top-left",
                "bg_color": "#101010",
                "accent_color": "#ffaa33",
                "style": "cinematic",
                "size_mode": "large",
                "overlay": "strong",
                "offset": "low",
                "artist": "artist name",
                "artist_case": "uppercase",
                "artist_font": "Helvetica Neue",
                "artist_color": "#ffffff",
                "artist_bold": "true",
                "artist_size": "72",
                "artist_x": "880",
                "artist_y": "1330",
                "title": "Title Card",
                "title_case": "normal",
                "title_font": "Helvetica Neue",
                "title_color": "#ffffff",
                "title_bold": True,
                "title_size": "140",
                "title_x": "900",
                "title_y": "1500",
                "subtitle": " deluxe ",
                "subtitle_case": "uppercase",
                "subtitle_font": "Inter Mono",
                "subtitle_color": "#d2d2d2",
                "subtitle_bold": False,
                "subtitle_size": "40",
                "subtitle_x": "910",
                "subtitle_y": "1610",
            },
            1800,
        )

        self.assertEqual("/tmp/cover.png", payload["artwork_path"])
        self.assertEqual(1.25, payload["zoom"])
        self.assertEqual("topleft", payload["layout"])
        self.assertEqual("cinematic", payload["options"]["style"])
        self.assertEqual("ARTIST NAME", payload["artist"])
        self.assertEqual("DELUXE", payload["subtitle"])
        self.assertEqual(3, len(payload["font_configs"]))
        self.assertEqual("ARTIST NAME", payload["font_configs"][0]["text"])
        self.assertTrue(payload["font_configs"][0]["bold"])
        self.assertEqual(880, payload["font_configs"][0]["x"])
        self.assertEqual(1610, payload["font_configs"][2]["y"])

    def test_compute_preview_layout_helpers_keep_image_visible(self):
        self.assertEqual(
            340,
            cover_preview_tools.compute_preview_stage_height(320),
        )
        self.assertEqual(
            (180, 360),
            cover_preview_tools.compute_preview_fit_size((1000, 2000), 220, 400, 360),
        )

    def test_read_artwork_meta_collects_size_and_dimensions(self):
        class FakeImage:
            size = (3000, 3000)

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

        details = cover_preview_tools.read_artwork_meta(
            "/tmp/cover.png",
            is_file_fn=lambda _path: True,
            getsize_fn=lambda _path: 1536,
            image_open=lambda _path: FakeImage(),
        )

        self.assertEqual("PNG", details["ext"])
        self.assertEqual("1.5 KB", details["size_text"])
        self.assertEqual((3000, 3000), details["dimensions"])

    def test_resolve_preview_refresh_and_zoom_states_cover_busy_open_missing_and_render(self):
        refresh_state = cover_preview_tools.resolve_preview_refresh_state(
            " /tmp/cover.png ",
            exists_fn=lambda path: path == "/tmp/cover.png",
        )
        busy = cover_preview_tools.resolve_preview_zoom_action(
            is_rendering=True,
            has_current_image=False,
            artwork_path="/tmp/cover.png",
            exists_fn=lambda _path: True,
        )
        open_state = cover_preview_tools.resolve_preview_zoom_action(
            is_rendering=False,
            has_current_image=True,
            artwork_path="/tmp/cover.png",
            exists_fn=lambda _path: True,
        )
        missing = cover_preview_tools.resolve_preview_zoom_action(
            is_rendering=False,
            has_current_image=False,
            artwork_path="/tmp/missing.png",
            exists_fn=lambda _path: False,
        )
        render_first = cover_preview_tools.resolve_preview_zoom_action(
            is_rendering=False,
            has_current_image=False,
            artwork_path="/tmp/cover.png",
            exists_fn=lambda path: path == "/tmp/cover.png",
        )

        self.assertEqual("/tmp/cover.png", refresh_state["artwork_path"])
        self.assertTrue(refresh_state["can_render"])
        self.assertEqual("busy", busy["action"])
        self.assertEqual("open", open_state["action"])
        self.assertEqual("missing", missing["action"])
        self.assertEqual("render_first", render_first["action"])


class DetailControllerToolsTests(unittest.TestCase):
    def test_build_detail_text_state_formats_tags_and_status(self):
        state = detail_controller_tools.build_detail_text_state(
            {
                "title": "Song A",
                "notes": "Idea",
                "tags": ["pop", "radio"],
                "instrumental": True,
                "status": "ready",
            }
        )

        self.assertEqual("Song A", state["title"])
        self.assertEqual("Idea", state["notes_text"])
        self.assertEqual("pop, radio", state["tags_text"])
        self.assertTrue(state["instrumental"])
        self.assertEqual("ready", state["status"])

    def test_populate_detail_view_sets_string_values(self):
        detail_vars = {"title": FakeVar(), "year": FakeVar()}

        detail_controller_tools.populate_detail_view(
            detail_vars,
            {"title": "Song A", "year": 2026},
        )

        self.assertEqual("Song A", detail_vars["title"].get())
        self.assertEqual("2026", detail_vars["year"].get())


class DetailViewToolsTests(unittest.TestCase):
    def test_build_detail_radar_state_handles_missing_audio_path(self):
        state = detail_view_tools.build_detail_radar_state(
            title="Song A",
            audio_path="/tmp/missing.wav",
            composer="Uwe",
            duration="3:21",
            year="2026",
            instrumental=False,
            status_text="Bereit",
            exists_fn=lambda _path: False,
        )

        self.assertEqual("Song A | Status: Bereit", state["headline"])
        self.assertIn("Audio fehlt: missing.wav", state["context_text"])
        self.assertIn("Komponist: Uwe", state["context_text"])
        self.assertIn("Jahr: 2026", state["context_text"])
        self.assertIn("Der gesetzte Audio-Pfad existiert nicht mehr", state["hint_text"])

    def test_build_detail_radar_state_handles_empty_record(self):
        state = detail_view_tools.build_detail_radar_state()

        self.assertEqual("Noch kein Werk geladen | Status: —", state["headline"])
        self.assertEqual("Audio: keines   •   Instrumental: Nein", state["context_text"])
        self.assertIn("Wähle ein bestehendes Werk", state["hint_text"])

    def test_build_detail_radar_state_highlights_missing_metadata_when_audio_is_linked(self):
        state = detail_view_tools.build_detail_radar_state(
            title="Song B",
            audio_path="/tmp/song-b.wav",
            composer="",
            duration="",
            year="2026",
            instrumental=True,
            status_text="In Arbeit",
            exists_fn=lambda _path: True,
        )

        self.assertIn("Audio: song-b.wav", state["context_text"])
        self.assertIn("Fehlt: Komponist und Dauer", state["context_text"])
        self.assertIn("Ergänze noch Komponist und Dauer", state["hint_text"])


@unittest.skipUnless(cover_tools.have_pillow(), "Pillow not available")
class LoudnessViewToolsTests(unittest.TestCase):
    def test_waveform_preview_output_path_uses_temp_preview_name(self):
        preview_path = loudness_view_tools.waveform_preview_output_path("/tmp/test-home")

        self.assertEqual("/tmp/test-home/.akm_temp/preview.png", preview_path)

    def test_render_waveform_preview_returns_resized_image(self):
        with tempfile.TemporaryDirectory() as tempdir:
            source_path = os.path.join(tempdir, "track.wav")
            with open(source_path, "wb") as handle:
                handle.write(b"audio")

            def fake_generate(_source, output_path, hex_color=None):
                _ = hex_color
                image = cover_tools.Image.new("RGB", (400, 100), color="black")
                image.save(output_path)
                return True

            preview_image, error = loudness_view_tools.render_waveform_preview(
                source_path,
                tempdir,
                "#ff9a3c",
                fake_generate,
                cover_tools.Image,
            )

            self.assertIsNone(error)
            self.assertEqual((800, 220), preview_image.size)


class LoudnessWorkflowTests(unittest.TestCase):
    def test_collect_importable_overview_audio_and_build_states(self):
        items = loudness_workflows.collect_importable_overview_audio(
            [
                {"title": "Song A", "audio_path": "/tmp/a.wav"},
                {"title": "Song B", "audio_path": "/tmp/missing.wav"},
            ],
            exists_fn=lambda path: path == "/tmp/a.wav",
        )
        loaded_state = loudness_workflows.build_loaded_files_state(
            ["/tmp/a.wav", "/tmp/b.wav"]
        )
        filtered_state = loudness_workflows.build_filtered_works_import_state(items)

        self.assertEqual([{"name": "Song A", "path": "/tmp/a.wav"}], items)
        self.assertEqual("2 Dateien geladen. Bereit für Analyse.", loaded_state["status_text"])
        self.assertEqual(["/tmp/a.wav"], filtered_state["paths"])
        self.assertEqual("1 Werke in Lautheit übernommen.", filtered_state["status_text"])

    def test_enrich_analysis_item_adds_match_fields(self):
        item = loudness_workflows.enrich_analysis_item(
            {
                "filename": "track.wav",
                "integrated_lufs": -16.0,
                "true_peak_dbtp": -2.0,
                "sample_peak_dbfs": -2.5,
                "ok": True,
            },
            -14.0,
            -1.0,
            FakeLoudnessModule,
        )

        self.assertEqual(2.0, item["gain_to_target_db"])
        self.assertEqual(0.0, item["predicted_true_peak_after_gain"])
        self.assertEqual("Peak Warnung", item["match_status"])
        self.assertEqual("Bereit", item["export_info"])

    def test_build_tree_row_formats_values_and_tags(self):
        row = loudness_workflows.build_tree_row(
            {
                "filename": "track.wav",
                "duration_display": "3:20",
                "integrated_lufs": -14.12,
                "true_peak_dbtp": -1.05,
                "sample_peak_dbfs": -1.4,
                "gain_to_target_db": 0.12,
                "predicted_true_peak_after_gain": -0.93,
                "match_status": "Peak Warnung",
                "export_info": "Bereit",
            }
        )

        self.assertEqual(("peak_warn",), row["tags"])
        self.assertEqual("track.wav", row["values"][0])
        self.assertEqual("-14.12", row["values"][2])
        self.assertEqual("Auto", row["values"][8])

    def test_export_result_item_marks_success_and_limiter_usage(self):
        result = loudness_workflows.export_result_item(
            {
                "filename": "track.wav",
                "path": "/tmp/track.wav",
                "gain_to_target_db": 1.5,
                "match_status": "Peak Warnung",
            },
            "/exports",
            -1.0,
            True,
            FakeLoudnessModule,
        )

        self.assertEqual(1, result["exported_increment"])
        self.assertEqual(0, result["warning_increment"])
        self.assertTrue(result["update"]["used_limiter"])
        self.assertEqual("Exportiert", result["update"]["export_info"])

    def test_apply_export_updates_merges_outputs_into_results(self):
        merged = loudness_workflows.apply_export_updates(
            [
                {"path": "/a.wav", "export_info": "Bereit"},
                {"path": "/b.wav", "export_info": "Bereit"},
            ],
            [
                {"path": "/b.wav", "export_info": "Exportiert", "output": "/out/b.wav"},
            ],
        )

        self.assertEqual("Bereit", merged[0]["export_info"])
        self.assertEqual("Exportiert", merged[1]["export_info"])
        self.assertEqual("/out/b.wav", merged[1]["exported_output"])


@unittest.skipUnless(cover_tools.have_pillow(), "Pillow not available")
class CoverToolsTests(unittest.TestCase):
    def test_resize_cover_canvas_returns_requested_size(self):
        image = cover_tools.Image.new("RGB", (400, 200), color="navy")

        resized = cover_tools.resize_cover_canvas(image, 300, 300)

        self.assertEqual((300, 300), resized.size)

    def test_generate_cover_variants_creates_expected_files(self):
        with tempfile.TemporaryDirectory() as tempdir:
            cover_path = os.path.join(tempdir, "cover.png")
            output_dir = os.path.join(tempdir, "out")
            os.makedirs(output_dir, exist_ok=True)

            image = cover_tools.Image.new("RGB", (800, 800), color="black")
            image.save(cover_path)

            created = cover_tools.generate_cover_variants(
                cover_path=cover_path,
                title="Test Release",
                artist="Test Artist",
                output_dir=output_dir,
                selected_layouts=["bottom", "topleft", "center"],
                selected_formats=["square"],
                options={
                    "style": "clean",
                    "size_mode": "medium",
                    "offset": "normal",
                    "overlay": "medium",
                },
            )

            self.assertEqual(3, len(created))
            self.assertTrue(created[0].endswith("cover_bottom_1x1.jpg"))
            self.assertTrue(created[1].endswith("cover_topleft_1x1.jpg"))
            self.assertTrue(created[2].endswith("cover_center_1x1.jpg"))
            for path in created:
                self.assertTrue(os.path.exists(path))


class ReleaseViewToolsTests(unittest.TestCase):
    def test_release_cover_output_dir_prefers_export_dir(self):
        output_dir = release_view_tools.release_cover_output_dir(
            "/tmp/export",
            "/tmp/source/cover.jpg",
        )

        self.assertEqual("/tmp/export/cover_previews", output_dir)

    def test_selected_release_cover_layouts_falls_back_to_defaults(self):
        selected = release_view_tools.selected_release_cover_layouts(
            {
                "bottom": False,
                "topleft": False,
                "center": False,
            }
        )

        self.assertEqual(["bottom", "topleft", "center"], selected)

    def test_build_release_track_row_label_and_status_text(self):
        label = release_view_tools.build_release_track_row_label(
            2,
            {
                "title": "Song",
                "duration": "3:12",
                "production": "Prod",
                "year": "2026",
                "source": "Werk",
            },
        )
        status = release_view_tools.build_release_status_text(
            4,
            True,
            False,
            True,
        )

        self.assertEqual("02. Song | 3:12 | Prod | 2026 | Werk", label)
        self.assertEqual(
            "4 Tracks im Release | Cover: Ja | Export-Ordner: Nein | Drag&Drop aktiv",
            status,
        )

    def test_release_view_helpers_build_preflight_flow_and_selection_hints(self):
        preflight = release_view_tools.build_release_preflight_text(
            2,
            "My Release",
            False,
            True,
            {"Werk": 1, "Datei→Werk": 1, "Datei": 0},
        )
        flow_hint = release_view_tools.build_release_flow_hint(
            2,
            "My Release",
            False,
            True,
            {"Werk": 1, "Datei→Werk": 1, "Datei": 0},
        )
        selection_hint = release_view_tools.build_release_selection_hint(3, (1,))

        self.assertEqual("Preflight: Cover fehlt", preflight)
        self.assertIn("automatisch auf Werke gemappt", flow_hint)
        self.assertIn("Cover", flow_hint)
        self.assertEqual(
            "1 Track ausgewählt. Verfügbar: Nach oben • Nach unten • Entfernen.",
            selection_hint,
        )


class AppRegressionTests(TemporaryStorageTestCase):
    def make_app_stub(self, records=None, mtime=None):
        app = SimpleNamespace()
        app.logs = []
        app.toasts = []
        app.state = FakeState(records=records, mtime=mtime)
        app.tasks = ImmediateTaskRunner()
        app.tab_system = SimpleNamespace(_instances={})
        app.get_built_tab = lambda tab_id: app.tab_system._instances.get(tab_id)
        app.cover_state_cache = {}
        app.release_state_cache = {}
        app.append_log = app.logs.append
        app.select_tab_by_id = lambda tab_id: setattr(app, "selected_tab", tab_id)
        app.status_text = lambda status: status
        app.get_overview_filter_state = lambda: (
            app.tab_system._instances["overview"].get_filter_state()
            if "overview" in app.tab_system._instances
            and hasattr(app.tab_system._instances["overview"], "get_filter_state")
            else {
                "search": (app.search_var.get() if getattr(app, "search_var", None) else ""),
                "filter": (app.status_filter_var.get() if getattr(app, "status_filter_var", None) else "all"),
                "sort": (app.sort_key_var.get() if getattr(app, "sort_key_var", None) else "title"),
                "desc": bool(app.sort_desc_var.get()) if getattr(app, "sort_desc_var", None) else False,
            }
        )
        app.get_detail_form_vars = lambda: (
            app.tab_system._instances["details"].get_form_vars()
            if "details" in app.tab_system._instances
            and hasattr(app.tab_system._instances["details"], "get_form_vars")
            else getattr(app, "detail_vars", {})
        )
        app.get_release_form_vars = lambda: (
            app.tab_system._instances["release"].get_form_state()
            if "release" in app.tab_system._instances
            and hasattr(app.tab_system._instances["release"], "get_form_state")
            else dict(app.release_state_cache)
        )
        return app

    def test_batch_controller_update_flow_applies_meta_progress_and_copy_label(self):
        app = self.make_app_stub()
        app.state.batch_queue = [
            {
                "title": "Song A",
                "composer": "Uwe",
                "duration": "3:11",
                "production": "FM",
                "year": "2026",
            }
        ]
        batch_view = FakeBatchView(copy_stage="duration")
        app.tab_system._instances["batch"] = batch_view

        controller = BatchController(app)
        controller.update_flow()

        self.assertEqual("Song A", app.current_title)
        self.assertEqual("Song A", batch_view.flow_state["title_text"])
        self.assertEqual(
            "Komponist: Uwe   Dauer: 3:11   Produktion: FM   Jahr: 2026",
            batch_view.flow_state["meta_text"],
        )
        self.assertEqual("1 / 1", batch_view.flow_state["progress_text"])
        self.assertEqual(100.0, batch_view.flow_state["progress_value"])
        self.assertEqual("Dauer kopieren", batch_view.flow_state["copy_button_label"])
        self.assertTrue(batch_view.flow_state["enabled"])

    def test_batch_tab_duration_button_stays_enabled_during_active_batch_flow(self):
        tab = BatchTab.__new__(BatchTab)
        tab.copy_stage = "duration"
        tab._batch_actions_enabled = True
        tab._focus_title_button = FakeConfigButton()
        tab._focus_duration_button = FakeConfigButton()

        BatchTab._update_focus_strip_buttons(tab)

        self.assertEqual("Titel", tab._focus_title_button.options["text"])
        self.assertEqual("Dauer •", tab._focus_duration_button.options["text"])
        self.assertEqual("normal", tab._focus_duration_button.options["state"])

    def test_batch_controller_flow_copy_updates_batch_view_copy_stage(self):
        app = self.make_app_stub()
        app.state.batch_queue = [
            {
                "title": "Song A",
                "duration": "3:11",
            }
        ]
        app.state.batch_index = 0
        app.clipboard_value = None
        app.clipboard_clear = lambda: setattr(app, "clipboard_value", "")
        app.clipboard_append = lambda value: setattr(app, "clipboard_value", value)
        batch_view = FakeBatchView(copy_stage="title")
        app.tab_system._instances["batch"] = batch_view

        controller = BatchController(app)
        controller.flow_copy()

        self.assertEqual("Song A", app.clipboard_value)
        self.assertEqual("duration", batch_view.copy_stage)
        self.assertEqual("Titel ✓", batch_view.copy_button_text)

    def test_batch_controller_flow_copy_duration_advances_to_next_track(self):
        app = self.make_app_stub()
        app.state.batch_queue = [
            {
                "title": "Song A",
                "duration": "3:11",
            },
            {
                "title": "Song B",
                "duration": "4:22",
            },
        ]
        app.state.batch_index = 0
        app.clipboard_value = None
        app.clipboard_clear = lambda: setattr(app, "clipboard_value", "")
        app.clipboard_append = lambda value: setattr(app, "clipboard_value", value)
        batch_view = FakeBatchView(copy_stage="duration")
        app.tab_system._instances["batch"] = batch_view

        controller = BatchController(app)
        controller.flow_copy()

        self.assertEqual("3:11", app.clipboard_value)
        self.assertEqual(1, app.state.batch_index)
        self.assertEqual("title", batch_view.copy_stage)
        self.assertEqual("Song B", batch_view.flow_state["title_text"])
        self.assertEqual("Titel kopieren", batch_view.flow_state["copy_button_label"])

    def test_batch_controller_flow_copy_duration_can_be_triggered_directly(self):
        app = self.make_app_stub()
        app.state.batch_queue = [
            {
                "title": "Song A",
                "duration": "3:11",
            },
            {
                "title": "Song B",
                "duration": "4:22",
            },
        ]
        app.state.batch_index = 0
        app.clipboard_value = None
        app.clipboard_clear = lambda: setattr(app, "clipboard_value", "")
        app.clipboard_append = lambda value: setattr(app, "clipboard_value", value)
        batch_view = FakeBatchView(copy_stage="title")
        app.tab_system._instances["batch"] = batch_view

        controller = BatchController(app)
        controller.flow_copy_duration()

        self.assertEqual("3:11", app.clipboard_value)
        self.assertEqual(1, app.state.batch_index)
        self.assertEqual("title", batch_view.copy_stage)
        self.assertEqual("Song B", batch_view.flow_state["title_text"])

    def test_batch_controller_update_flow_backfills_missing_duration_from_audio(self):
        app = self.make_app_stub()
        app.state.batch_queue = [
            {
                "title": "Song A",
                "duration": "",
                "audio_path": "/tmp/song-a.wav",
            }
        ]
        batch_view = FakeBatchView(copy_stage="title")
        app.tab_system._instances["batch"] = batch_view

        controller = BatchController(app)
        with mock.patch.object(loudness_tools, "probe_duration", return_value=191.0) as probe_duration:
            with mock.patch.object(akm_core, "update_entry", return_value=(True, {"duration": "3:11"})) as update_entry:
                controller.update_flow()

        probe_duration.assert_called_once_with("/tmp/song-a.wav")
        update_entry.assert_called_once_with("Song A", {"duration": "3:11"})
        self.assertEqual("3:11", app.state.batch_queue[0]["duration"])
        self.assertIn("Dauer: 3:11", batch_view.flow_state["meta_text"])
        self.assertTrue(app.state.invalidated)

    def test_batch_controller_flow_copy_duration_probes_audio_when_duration_missing(self):
        app = self.make_app_stub()
        app.state.batch_queue = [
            {
                "title": "Song A",
                "duration": "",
                "audio_path": "/tmp/song-a.wav",
            },
            {
                "title": "Song B",
                "duration": "4:22",
            },
        ]
        app.state.batch_index = 0
        app.clipboard_value = None
        app.clipboard_clear = lambda: setattr(app, "clipboard_value", "")
        app.clipboard_append = lambda value: setattr(app, "clipboard_value", value)
        batch_view = FakeBatchView(copy_stage="duration")
        app.tab_system._instances["batch"] = batch_view

        controller = BatchController(app)
        with mock.patch.object(loudness_tools, "probe_duration", return_value=191.0) as probe_duration:
            with mock.patch.object(akm_core, "update_entry", return_value=(True, {"duration": "3:11"})) as update_entry:
                controller.flow_copy_duration()

        probe_duration.assert_called_once_with("/tmp/song-a.wav")
        update_entry.assert_called_once_with("Song A", {"duration": "3:11"})
        self.assertEqual("3:11", app.clipboard_value)
        self.assertEqual(1, app.state.batch_index)
        self.assertEqual("Song B", batch_view.flow_state["title_text"])

    def test_project_controller_import_done_logs_summary_and_refreshes_overview(self):
        app = self.make_app_stub()
        app.overview_ctrl = SimpleNamespace(
            refresh_list=lambda: setattr(app, "overview_refreshed", True)
        )
        app.batch_ctrl = SimpleNamespace(
            reload_flow_data=lambda preferred_index=None: setattr(app, "batch_refreshed", preferred_index)
        )
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = app.toasts.append

        controller._on_import_done(
            [
                {
                    "title": "Song A",
                    "action": "added",
                    "duration": "3:11",
                    "composer": "Uwe",
                },
                {
                    "title": "Song B",
                    "action": "updated",
                    "duration": "",
                    "composer": "",
                },
            ]
        )

        self.assertEqual("Excel-Import abgeschlossen: 2 Einträge", app.logs[0])
        self.assertEqual("  + Song A (3:11 | Uwe)", app.logs[1])
        self.assertEqual("  ~ Song B", app.logs[2])
        self.assertEqual("Excel-Import abgeschlossen: 2 Einträge", app.toasts[0])
        self.assertTrue(app.state.invalidated)
        self.assertTrue(app.overview_refreshed)
        self.assertEqual(0, app.batch_refreshed)

    def test_project_controller_add_entry_uses_remembered_detail_defaults(self):
        app = self.make_app_stub()
        app.overview_ctrl = SimpleNamespace(
            _on_g_done=lambda result, message: setattr(app, "done_payload", (result, message))
        )
        controller = ProjectController(app)
        akm_core.remember_detail_memory(
            {
                "composer": "Uwe",
                "production": "FM",
                "year": "2026",
            }
        )

        with mock.patch(
            "app_controllers.project_controller.akm_core.add_entry",
            return_value=(True, {"title": "Song A"}),
        ) as add_entry:
            controller.add_entry("Song A")

        add_entry.assert_called_once_with(
            "Song A",
            composer="Uwe",
            production="FM",
            year="2026",
        )
        self.assertEqual("'Song A' angelegt", app.done_payload[1])

    def test_project_controller_save_project_returns_false_when_dialog_cancelled(self):
        app = self.make_app_stub()
        app.cover_tab = SimpleNamespace(get_state=lambda: {})
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=None)
        app.wait_window = lambda dialog: None

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmSaveDialog",
            return_value=fake_dialog,
        ):
            saved = controller.save_project()

        self.assertFalse(saved)
        self.assertIn("Speichervorgang abgebrochen.", app.logs)

    def test_project_controller_save_project_uses_cached_release_state_without_built_release_tab(self):
        app = self.make_app_stub()
        app.release_state_cache = {
            "artist": "Loaded Artist",
            "type": "EP",
            "cover_path": "/tmp/cover.png",
        }
        app.cover_tab = SimpleNamespace(get_state=lambda: {})
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=os.path.join(self.tempdir.name, "demo.akm"))
        app.wait_window = lambda dialog: None

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmSaveDialog",
            return_value=fake_dialog,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.PROJECTS_DIR",
            self.tempdir.name,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.save_project",
        ) as save_project:
            saved = controller.save_project()

        self.assertTrue(saved)
        release_state = save_project.call_args.args[3]
        self.assertEqual(app.release_state_cache, release_state["vars"])
        self.assertEqual([], release_state["tracks"])

    def test_project_controller_save_project_uses_cached_cover_state_without_built_cover_tab(self):
        app = self.make_app_stub()
        app.cover_state_cache = {
            "title": "Loaded Cover Title",
            "layout": "center",
            "accent_color": "#ffaa00",
        }
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=os.path.join(self.tempdir.name, "demo.akm"))
        app.wait_window = lambda dialog: None

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmSaveDialog",
            return_value=fake_dialog,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.PROJECTS_DIR",
            self.tempdir.name,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.save_project",
        ) as save_project:
            saved = controller.save_project()

        self.assertTrue(saved)
        cover_state = save_project.call_args.args[2]
        self.assertEqual(app.cover_state_cache, cover_state)

    def test_project_controller_save_project_prefers_built_cover_tab_state_over_cache(self):
        app = self.make_app_stub()
        app.cover_state_cache = {"title": "Stale Cached Title"}
        app.tab_system._instances["cover"] = SimpleNamespace(
            get_state=lambda: {"title": "Live Cover Title", "layout": "manual"}
        )
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=os.path.join(self.tempdir.name, "demo.akm"))
        app.wait_window = lambda dialog: None

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmSaveDialog",
            return_value=fake_dialog,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.PROJECTS_DIR",
            self.tempdir.name,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.save_project",
        ) as save_project:
            saved = controller.save_project()

        self.assertTrue(saved)
        cover_state = save_project.call_args.args[2]
        self.assertEqual({"title": "Live Cover Title", "layout": "manual"}, cover_state)
        self.assertEqual(cover_state, app.cover_state_cache)

    def test_project_controller_save_project_prefers_built_release_tab_state_over_cache(self):
        app = self.make_app_stub()
        app.release_state_cache = {"artist": "Stale Artist", "type": "Single"}
        app.tab_system._instances["release"] = SimpleNamespace(
            get_form_state=lambda: {"artist": "Live Artist", "type": "Album"}
        )
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=os.path.join(self.tempdir.name, "demo.akm"))
        app.wait_window = lambda dialog: None

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmSaveDialog",
            return_value=fake_dialog,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.PROJECTS_DIR",
            self.tempdir.name,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.save_project",
        ) as save_project:
            saved = controller.save_project()

        self.assertTrue(saved)
        release_state = save_project.call_args.args[3]
        self.assertEqual({"artist": "Live Artist", "type": "Album"}, release_state["vars"])
        self.assertEqual(release_state["vars"], app.release_state_cache)

    def test_save_project_serializes_track_records_in_catalog_and_release_tracks(self):
        project_path = os.path.join(self.tempdir.name, "demo.akm")

        akm_core.save_project(
            project_path,
            [
                TrackRecord(
                    title="Song A",
                    composer="Uwe",
                    tags=["single"],
                )
            ],
            cover_state={"title": "Cover"},
            release_state={
                "vars": {"artist": "Live Artist"},
                "tracks": [
                    TrackRecord(
                        title="Song A",
                        audio_path="/tmp/song.wav",
                    )
                ],
            },
        )

        with open(project_path, "r", encoding="utf-8") as handle:
            bundle = json.load(handle)

        self.assertEqual("Song A", bundle["data"][0]["title"])
        self.assertEqual("Uwe", bundle["data"][0]["composer"])
        self.assertEqual(["single"], bundle["data"][0]["tags"])
        self.assertEqual("Song A", bundle["release"]["tracks"][0]["title"])
        self.assertEqual("/tmp/song.wav", bundle["release"]["tracks"][0]["audio_path"])

    def test_project_controller_load_project_caches_cover_state_without_built_cover_tab(self):
        app = self.make_app_stub()
        app.overview_ctrl = SimpleNamespace(
            refresh_list=lambda: setattr(app, "overview_refreshed", True)
        )
        app.release_ctrl = SimpleNamespace(
            refresh_view=lambda force=False: setattr(app, "release_refresh_force", force)
        )
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=os.path.join(self.tempdir.name, "demo.akm"))
        app.wait_window = lambda dialog: None
        bundle = {
            "data": [{"title": "Song A"}],
            "cover": {"title": "Loaded Cover Title", "layout": "center"},
        }

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmLoadDialog",
            return_value=fake_dialog,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.PROJECTS_DIR",
            self.tempdir.name,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.load_project",
            return_value=bundle,
        ):
            controller.load_project_dialog()

        self.assertEqual(bundle["cover"], app.cover_state_cache)
        self.assertTrue(app.overview_refreshed)

    def test_project_controller_load_project_updates_built_cover_tab_state(self):
        app = self.make_app_stub()
        app.overview_ctrl = SimpleNamespace(
            refresh_list=lambda: setattr(app, "overview_refreshed", True)
        )
        app.release_ctrl = SimpleNamespace(
            refresh_view=lambda force=False: setattr(app, "release_refresh_force", force)
        )
        app.tab_system._instances["cover"] = SimpleNamespace(
            set_state=lambda state: setattr(app, "loaded_cover_state", state)
        )
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=os.path.join(self.tempdir.name, "demo.akm"))
        app.wait_window = lambda dialog: None
        bundle = {
            "data": [{"title": "Song A"}],
            "cover": {"title": "Loaded Cover Title", "layout": "center"},
        }

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmLoadDialog",
            return_value=fake_dialog,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.PROJECTS_DIR",
            self.tempdir.name,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.load_project",
            return_value=bundle,
        ):
            controller.load_project_dialog()

        self.assertEqual(bundle["cover"], app.cover_state_cache)
        self.assertEqual(bundle["cover"], app.loaded_cover_state)

    def test_project_controller_load_project_caches_release_state_without_built_release_tab(self):
        app = self.make_app_stub()
        app.overview_ctrl = SimpleNamespace(
            refresh_list=lambda: setattr(app, "overview_refreshed", True)
        )
        app.release_ctrl = SimpleNamespace(
            refresh_view=lambda force=False: setattr(app, "release_refresh_force", force)
        )
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=os.path.join(self.tempdir.name, "demo.akm"))
        app.wait_window = lambda dialog: None
        bundle = {
            "data": [{"title": "Song A"}],
            "release": {
                "vars": {"artist": "Loaded Artist", "cover_path": "/tmp/cover.png"},
                "tracks": [{"title": "Song A", "audio_path": "/tmp/song.wav"}],
            },
        }

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmLoadDialog",
            return_value=fake_dialog,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.PROJECTS_DIR",
            self.tempdir.name,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.load_project",
            return_value=bundle,
        ):
            controller.load_project_dialog()

        self.assertEqual(bundle["release"]["vars"], app.release_state_cache)
        self.assertEqual(bundle["release"]["tracks"], app.state.release_tracks)
        self.assertTrue(app.overview_refreshed)
        self.assertTrue(app.release_refresh_force)

    def test_project_controller_load_project_updates_built_release_tab_state(self):
        app = self.make_app_stub()
        app.overview_ctrl = SimpleNamespace(
            refresh_list=lambda: setattr(app, "overview_refreshed", True)
        )
        app.release_ctrl = SimpleNamespace(
            refresh_view=lambda force=False: setattr(app, "release_refresh_force", force)
        )
        app.tab_system._instances["release"] = SimpleNamespace(
            set_form_state=lambda state: setattr(app, "loaded_release_state", state)
        )
        controller = ProjectController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        fake_dialog = SimpleNamespace(result=os.path.join(self.tempdir.name, "demo.akm"))
        app.wait_window = lambda dialog: None
        bundle = {
            "data": [{"title": "Song A"}],
            "release": {
                "vars": {"artist": "Loaded Artist", "cover_path": "/tmp/cover.png"},
                "tracks": [{"title": "Song A", "audio_path": "/tmp/song.wav"}],
            },
        }

        with mock.patch(
            "app_controllers.project_controller.ui_patterns.AkmLoadDialog",
            return_value=fake_dialog,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.PROJECTS_DIR",
            self.tempdir.name,
        ), mock.patch(
            "app_controllers.project_controller.akm_core.load_project",
            return_value=bundle,
        ):
            controller.load_project_dialog()

        self.assertEqual(bundle["release"]["vars"], app.release_state_cache)
        self.assertEqual(bundle["release"]["vars"], app.loaded_release_state)
        self.assertEqual(bundle["release"]["tracks"], app.state.release_tracks)
        self.assertTrue(app.release_refresh_force)

    def test_on_closing_keeps_window_open_when_save_is_cancelled(self):
        app = SimpleNamespace(
            save_project=lambda: False,
            destroy=lambda: setattr(app, "destroyed", True),
        )
        app.destroyed = False

        with mock.patch.object(akm_app, "messagebox", FakeMessagebox(True)):
            akm_app.AKMApp.on_closing(app)

        self.assertFalse(app.destroyed)

    def test_on_closing_destroys_window_after_successful_save(self):
        app = SimpleNamespace(
            save_project=lambda: True,
            destroy=lambda: setattr(app, "destroyed", True),
        )
        app.destroyed = False

        with mock.patch.object(akm_app, "messagebox", FakeMessagebox(True)):
            akm_app.AKMApp.on_closing(app)

        self.assertTrue(app.destroyed)

    def test_quit_shortcut_routes_through_on_closing(self):
        app = SimpleNamespace(
            on_closing=lambda: setattr(app, "closing_called", True)
        )
        app.closing_called = False

        result = akm_app.AKMApp._on_quit_shortcut(app)

        self.assertTrue(app.closing_called)
        self.assertEqual("break", result)

    def test_refresh_all_tabs_includes_details_and_built_cover_refresh(self):
        app = SimpleNamespace(
            search_var=FakeVar("Song"),
            status_filter_var=FakeVar("ready"),
            sort_key_var=FakeVar("year"),
            sort_desc_var=FakeVar(True),
            state=FakeState(mtime=42),
            overview_ctrl=SimpleNamespace(
                refresh_list=lambda: setattr(app, "overview_refreshed", True),
                refresh_dashboard=lambda: setattr(app, "dashboard_refreshed", True),
            ),
            details_ctrl=SimpleNamespace(
                refresh_view=lambda: setattr(app, "details_refreshed", True)
            ),
            batch_ctrl=SimpleNamespace(
                reload_flow_data=lambda: setattr(app, "batch_refreshed", True)
            ),
            release_ctrl=SimpleNamespace(
                refresh_view=lambda: setattr(app, "release_refreshed", True)
            ),
            tab_system=SimpleNamespace(
                _instances={
                    "cover": SimpleNamespace(
                        refresh_view=lambda: setattr(app, "cover_refreshed", True)
                    )
                }
            ),
            _last_overview_refresh={
                "search": None,
                "filter": None,
                "sort": None,
                "desc": None,
                "mtime": None,
            },
            _last_dashboard_refresh={"mtime": None},
            _last_batch_refresh={"mtime": None},
        )
        app.overview_refreshed = False
        app.dashboard_refreshed = False
        app.details_refreshed = False
        app.batch_refreshed = False
        app.release_refreshed = False
        app.cover_refreshed = False

        akm_app.AKMApp.refresh_all_tabs(app)

        self.assertTrue(app.overview_refreshed)
        self.assertTrue(app.dashboard_refreshed)
        self.assertTrue(app.details_refreshed)
        self.assertTrue(app.batch_refreshed)
        self.assertTrue(app.release_refreshed)
        self.assertTrue(app.cover_refreshed)
        self.assertEqual(
            {
                "search": "song",
                "filter": "ready",
                "sort": "year",
                "desc": True,
                "mtime": 42,
            },
            app._last_overview_refresh,
        )
        self.assertEqual(42, app._last_dashboard_refresh["mtime"])
        self.assertEqual(42, app._last_batch_refresh["mtime"])

    def test_open_loudness_tab_updates_built_loudness_view(self):
        app = self.make_app_stub()
        loudness_view = FakeLoudnessView()
        app.loudness_tab = loudness_view

        with mock.patch(
            "akm_app.assistant_tools.build_loudness_tab_open_state",
            return_value={
                "status_text": "Lautheit bereit",
                "hint_text": "Ziehe Dateien hier hinein.",
                "log_message": "Lautheit geöffnet",
            },
        ):
            akm_app.AKMApp.open_loudness_tab(app)

        self.assertEqual("loudness", app.selected_tab)
        self.assertEqual("Lautheit bereit", loudness_view.status_text)
        self.assertEqual("Ziehe Dateien hier hinein.", loudness_view.hint_text)
        self.assertIn("Lautheit geöffnet", app.logs)

    def test_overview_controller_refresh_list_clears_stale_ui_when_no_records(self):
        app = self.make_app_stub(records=[], mtime=None)
        app.state.filtered_records = [{"title": "Veraltet"}]
        overview_view = FakeOverviewView(
            filter_state={"search": "", "filter": "all", "sort": "title", "desc": False}
        )
        overview_view.list_items = ["Stale Row"]
        app.tab_system._instances["overview"] = overview_view

        controller = OverviewController(app)
        controller.refresh_list()

        self.assertEqual([], app.state.filtered_records)
        self.assertEqual([], overview_view.list_items)
        self.assertEqual(
            "0 Treffer   •   Sortierung: Titel aufsteigend",
            overview_view.summary_text,
        )
        self.assertEqual("all  0", overview_view.filter_chip_texts["all"])
        self.assertTrue(overview_view.filter_chip_selected["all"])

    def test_overview_controller_refresh_list_does_not_cache_without_built_view(self):
        app = self.make_app_stub(
            records=[{"title": "Song A", "status": "ready"}],
            mtime=42,
        )
        controller = OverviewController(app)

        controller.refresh_list()

        self.assertEqual(1, len(app.state.filtered_records))
        self.assertFalse(hasattr(controller, "_last_refresh_params"))

        overview_view = FakeOverviewView()
        app.tab_system._instances["overview"] = overview_view

        controller.refresh_list()

        self.assertEqual(1, len(overview_view.list_items))
        self.assertTrue(hasattr(controller, "_last_refresh_params"))

    def test_overview_controller_refresh_dashboard_updates_dashboard_view(self):
        app = self.make_app_stub(
            records=[
                {"title": "Song A", "status": "ready", "production": "P", "notes": "n", "instrumental": False},
                {"title": "Song B", "status": "confirmed", "production": "", "notes": "", "instrumental": True},
            ]
        )
        dashboard_view = FakeDashboardView()
        app.tab_system._instances["dashboard"] = dashboard_view

        controller = OverviewController(app)
        controller.refresh_dashboard()

        self.assertEqual(2, dashboard_view.stats["total"])
        self.assertIn("2 Werke", dashboard_view.status_text)
        self.assertIn("Mit Produktion: 1", dashboard_view.meta_text)
        self.assertEqual("ready  1", dashboard_view.chip_labels["ready"])

    def test_overview_controller_open_with_filter_builds_overview_view_state(self):
        app = self.make_app_stub()
        overview_view = FakeOverviewView()
        app.tab_system._instances["overview"] = overview_view
        app.overview_tab = overview_view
        app.overview_ctrl = OverviewController(app)
        app.overview_refreshed = False
        app.overview_ctrl.refresh_list = lambda: setattr(app, "overview_refreshed", True)

        app.overview_ctrl.open_with_filter("confirmed")

        self.assertEqual("overview", app.selected_tab)
        self.assertEqual("confirmed", overview_view.filter_state["filter"])
        self.assertTrue(app.overview_refreshed)

    def test_overview_controller_load_selected_into_details_uses_details_view(self):
        record = {
            "title": "Song A",
            "status": "ready",
            "duration": "3:11",
            "composer": "Uwe",
            "production": "FM",
            "year": "2026",
            "audio_path": "/tmp/song.wav",
            "notes": "Notiz",
            "tags": ["tag1", "tag2"],
            "instrumental": True,
        }
        app = self.make_app_stub(records=[record])
        app.state.filtered_records = [record]
        app.get_detail_form_vars = lambda: details_view.get_form_vars()
        app.details_ctrl = SimpleNamespace(
            set_status_chip=lambda status: setattr(app, "detail_status", status)
        )
        overview_view = FakeOverviewView(selection=(0,))
        details_view = FakeDetailsView(
            form_vars={
                "title": FakeVar(""),
                "duration": FakeVar(""),
                "composer": FakeVar(""),
                "production": FakeVar(""),
                "year": FakeVar(""),
                "audio_path": FakeVar(""),
            }
        )
        app.tab_system._instances["overview"] = overview_view
        app.tab_system._instances["details"] = details_view

        controller = OverviewController(app)
        controller.load_selected_into_details()

        self.assertEqual("Song A", details_view.form_vars["title"].get())
        self.assertEqual("Notiz", details_view.notes_text)
        self.assertEqual("tag1, tag2", details_view.tags_text)
        self.assertTrue(details_view.instrumental)
        self.assertEqual("ready", app.detail_status)
        self.assertEqual("details", app.selected_tab)

    def test_details_controller_save_details_updates_loaded_record(self):
        self.save_entries(
            [
                {
                    "title": "Song A",
                    "status": "in_progress",
                    "date": "2026-01-01",
                    "last_change": "2026-01-01",
                    "duration": "",
                    "composer": "",
                    "production": "",
                    "year": "",
                    "instrumental": False,
                    "notes": "old A",
                    "audio_path": "",
                    "tags": [],
                },
                {
                    "title": "Song B",
                    "status": "in_progress",
                    "date": "2026-01-01",
                    "last_change": "2026-01-01",
                    "duration": "",
                    "composer": "",
                    "production": "",
                    "year": "",
                    "instrumental": False,
                    "notes": "old B",
                    "audio_path": "",
                    "tags": [],
                },
            ]
        )

        app = self.make_app_stub()
        app.detail_original_title = "Song A"
        app.current_detail_status = "ready"
        details_view = FakeDetailsView(
            form_vars={
            "title": FakeVar("Song A"),
            "duration": FakeVar(""),
            "composer": FakeVar(""),
            "production": FakeVar(""),
            "year": FakeVar(""),
            "audio_path": FakeVar(""),
            }
        )
        details_view.set_notes_text("updated A")
        details_view.set_tags_text("")
        details_view.set_instrumental(False)
        app.tab_system._instances["details"] = details_view
        app.overview_ctrl = SimpleNamespace(
            _on_g_done=lambda result, message: setattr(app, "saved_payload", (result, message))
        )

        controller = DetailsController(app)
        controller.save_details()

        entries = {item["title"]: item for item in self.load_entries()}
        self.assertEqual("updated A", entries["Song A"]["notes"])
        self.assertEqual("old B", entries["Song B"]["notes"])
        self.assertEqual(
            {
                "composer": "",
                "production": "",
                "year": "",
            },
            akm_core.get_detail_memory(),
        )
        self.assertTrue(app.saved_payload[0][0])
        self.assertEqual("Gespeichert", app.saved_payload[1])

    def test_details_controller_clear_details_form_restores_memory_defaults(self):
        akm_core.remember_detail_memory(
            {
                "composer": "Uwe",
                "production": "FM",
                "year": "2026",
            }
        )

        app = self.make_app_stub()
        app.detail_original_title = "Song A"
        details_view = FakeDetailsView(
            form_vars={
                "title": FakeVar("Song A"),
                "duration": FakeVar("3:11"),
                "composer": FakeVar("Other"),
                "production": FakeVar("Other Prod"),
                "year": FakeVar("2025"),
                "audio_path": FakeVar("/tmp/song.wav"),
            }
        )
        details_view.set_notes_text("note")
        details_view.set_tags_text("tag")
        details_view.set_instrumental(True)
        app.tab_system._instances["details"] = details_view

        controller = DetailsController(app)
        controller.clear_details_form()

        self.assertIsNone(app.detail_original_title)
        self.assertEqual("", details_view.form_vars["title"].get())
        self.assertEqual("", details_view.form_vars["duration"].get())
        self.assertEqual("", details_view.form_vars["audio_path"].get())
        self.assertEqual("Uwe", details_view.form_vars["composer"].get())
        self.assertEqual("FM", details_view.form_vars["production"].get())
        self.assertEqual("2026", details_view.form_vars["year"].get())
        self.assertEqual("", details_view.notes_text)
        self.assertEqual("", details_view.tags_text)
        self.assertFalse(details_view.instrumental)
        self.assertEqual("in_progress", app.current_detail_status)

    def test_details_controller_refresh_titles_populates_after_late_tab_build(self):
        app = self.make_app_stub(
            records=[
                {"title": "Song B"},
                {"title": "Song A"},
            ]
        )
        controller = DetailsController(app)

        controller.refresh_titles()

        self.assertIsNone(controller._title_signature)

        details_view = FakeDetailsView(form_vars={"title": FakeVar("")})
        app.tab_system._instances["details"] = details_view

        controller.refresh_titles()

        self.assertEqual(["Song A", "Song B"], details_view.title_values)
        self.assertEqual((app.state._get_data_mtime(), 2), controller._title_signature)

    def test_details_controller_set_status_chip_updates_public_state(self):
        app = self.make_app_stub()
        app.current_detail_status = "in_progress"
        details_view = FakeDetailsView()
        app.tab_system._instances["details"] = details_view
        app.status_text = lambda status: f"label:{status}"

        controller = DetailsController(app)
        controller.set_status_chip("ready")

        self.assertEqual("ready", app.current_detail_status)
        self.assertEqual("ready", details_view.status_key)
        self.assertEqual("label:ready", details_view.status_text)

    def test_loudness_controller_import_filtered_works_updates_state_and_labels(self):
        app = self.make_app_stub()
        app.state.filtered_records = [
            {"title": "Song A", "audio_path": "/tmp/a.wav"},
            {"title": "Song B", "audio_path": "/tmp/missing.wav"},
        ]
        loudness_view = FakeLoudnessView()
        app.tab_system._instances["loudness"] = loudness_view

        controller = LoudnessController(app)
        controller._pop_l_tree = lambda: setattr(app, "loudness_tree_cleared", True)
        controller.log = app.logs.append

        with mock.patch(
            "app_workflows.loudness_workflows.os.path.exists",
            side_effect=lambda path: path == "/tmp/a.wav",
        ):
            controller.import_filtered_works()

        self.assertEqual(["/tmp/a.wav"], app.state.loudness_files)
        self.assertEqual([], app.state.loudness_results)
        self.assertTrue(app.loudness_tree_cleared)
        self.assertEqual("1 Werke in Lautheit übernommen.", loudness_view.status_text)
        self.assertEqual("loudness", app.selected_tab)
        self.assertIn("Aus Werken übernommen: 1 Dateien", app.logs)

    def test_loudness_controller_analyze_files_reads_settings_from_loudness_view(self):
        app = self.make_app_stub()
        app.state.loudness_files = ["/tmp/song.wav"]
        loudness_view = FakeLoudnessView(target_text="-12,5", peak_text="-0.5")
        app.tab_system._instances["loudness"] = loudness_view

        controller = LoudnessController(app)
        controller._pop_l_tree = lambda: setattr(app, "loudness_tree_reloaded", True)
        controller.log = app.logs.append

        with mock.patch(
            "app_controllers.loudness_controller.loudness_tools.analyze_full_track",
            return_value={
                "filename": "song.wav",
                "path": "/tmp/song.wav",
                "ok": True,
                "duration_display": "3:00",
                "integrated_lufs": -10.0,
                "true_peak_dbtp": -1.2,
                "sample_peak_dbfs": -1.4,
            },
        ):
            controller.analyze_files()

        self.assertTrue(app.loudness_tree_reloaded)
        self.assertEqual(-12.5, app.state.loudness_results[0]["target_lufs"])
        self.assertEqual(-2.5, app.state.loudness_results[0]["gain_to_target_db"])
        self.assertEqual("Analyse abgeschlossen: 1 Dateien.", app.logs[-1])

    def test_loudness_controller_analyze_files_invalid_settings_show_error(self):
        app = self.make_app_stub()
        app.state.loudness_files = ["/tmp/song.wav"]
        loudness_view = FakeLoudnessView(target_text="nope", peak_text="-0.5")
        app.tab_system._instances["loudness"] = loudness_view

        controller = LoudnessController(app)

        with mock.patch("app_controllers.loudness_controller.messagebox") as messagebox_mock:
            controller.analyze_files()

        messagebox_mock.showerror.assert_called_once_with(
            "Eingabefehler",
            "LUFS oder Peak-Wert ist kein gültiges Zahlenformat.",
        )
        self.assertEqual([], app.state.loudness_results)

    def test_release_controller_handle_drop_deduplicates_and_matches_titles(self):
        matched_path = os.path.join(self.tempdir.name, "01 - Intro_matched.wav")
        with open(matched_path, "wb") as handle:
            handle.write(b"audio")

        app = self.make_app_stub(
            records=[
                {
                    "title": "Intro",
                    "duration": "1:11",
                    "composer": "Uwe",
                    "production": "Prod",
                    "year": "2026",
                    "notes": "Ready for AKM",
                    "audio_path": "",
                }
            ]
        )
        release_view = FakeReleaseView()
        app.tab_system._instances["release"] = release_view
        app.tasks.parse_dnd_files = lambda _data: [matched_path, matched_path]

        controller = ReleaseController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        controller.handle_drop(SimpleNamespace(data="ignored"))

        self.assertEqual(1, len(app.state.release_tracks))
        self.assertEqual("Datei→Werk", app.state.release_tracks[0]["source"])
        self.assertEqual("Uwe", app.state.release_tracks[0]["composer"])
        self.assertEqual("Ready for AKM", app.state.release_tracks[0]["notes"])
        self.assertEqual("01. Intro | 1:11 | Prod | 2026 | Datei→Werk", release_view.track_labels[0])
        self.assertIn("Release DnD: 1 Tracks hinzugefügt.", app.logs)
        self.assertEqual("1 TRACKS HINZUGEFÜGT", app.toasts[0][0])

    def test_release_controller_import_release_to_batch_updates_catalog_and_opens_batch(self):
        app = self.make_app_stub()
        app.overview_ctrl = SimpleNamespace(
            refresh_list=lambda: setattr(app, "overview_refreshed", True)
        )
        app.batch_ctrl = SimpleNamespace(
            reload_flow_data=lambda preferred_index=None: setattr(app, "batch_refreshed", preferred_index)
        )
        app.state.release_tracks = [
            {
                "title": "Intro",
                "duration": "1:11",
                "composer": "Uwe",
                "production": "Prod",
                "year": "2026",
                "notes": "Ready for AKM",
                "audio_path": "/tmp/intro.wav",
                "source": "Werk",
            }
        ]

        controller = ReleaseController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        with mock.patch(
            "app_controllers.release_controller.akm_core.import_tracks",
            return_value=[
                {
                    "title": "Intro",
                    "duration": "1:11",
                    "composer": "Uwe",
                    "action": "added",
                }
            ],
        ) as import_tracks:
            controller.import_release_to_batch(open_batch=True)

        import_tracks.assert_called_once_with(
            [
                {
                    "title": "Intro",
                    "duration": "1:11",
                    "composer": "Uwe",
                    "production": "Prod",
                    "year": "2026",
                    "notes": "Ready for AKM",
                    "audio_path": "/tmp/intro.wav",
                }
            ]
        )
        self.assertEqual("Release -> AKM abgeschlossen: 1 Einträge", app.logs[0])
        self.assertEqual("  + Intro (1:11 | Uwe)", app.logs[1])
        self.assertEqual("Release -> AKM abgeschlossen: 1 Einträge", app.toasts[0][0])
        self.assertTrue(app.state.invalidated)
        self.assertTrue(app.overview_refreshed)
        self.assertEqual(0, app.batch_refreshed)
        self.assertEqual("batch", app.selected_tab)

    def test_release_controller_build_export_triggers_release_handoff_on_success(self):
        app = self.make_app_stub()
        app.release_state_cache = {"title": "My Release"}
        app.state.release_tracks = [{"title": "Intro", "audio_path": "/tmp/intro.wav", "source": "Werk"}]

        controller = ReleaseController(app)
        controller.log = app.logs.append
        controller.toast = lambda message, **kwargs: app.toasts.append((message, kwargs))

        with mock.patch(
            "app_controllers.release_controller.release_workflows.start_distro_export",
            return_value=(True, "Export fertig"),
        ) as export_mock, mock.patch.object(
            controller,
            "import_release_to_batch",
        ) as import_release:
            controller.build_export()

        export_mock.assert_called_once_with({"title": "My Release"}, app.state.release_tracks)
        import_release.assert_called_once_with(open_batch=True)
        self.assertIn("Export fertig", app.logs)
        self.assertEqual("EXPORT FERTIG", app.toasts[0][0])

    def test_release_controller_refresh_view_waits_for_built_release_tab(self):
        app = self.make_app_stub()
        app.state.release_tracks = [{"title": "Song A", "audio_path": "/tmp/song.wav", "source": "Datei"}]
        app.release_state_cache = {"title": "My Release", "cover_path": "/tmp/cover.png", "export_dir": "/tmp/export"}

        controller = ReleaseController(app)
        controller.refresh_view()

        self.assertIsNone(controller._last_view_signature)

        release_view = FakeReleaseView(form_state=app.release_state_cache)
        app.tab_system._instances["release"] = release_view

        controller.refresh_view()

        self.assertEqual(1, len(release_view.track_labels))
        self.assertEqual("Preflight: 1 ohne Werk-Match", release_view.preflight_text)
        self.assertIn("Export erstellt Paket", release_view.flow_hint)
        self.assertIsNotNone(controller._last_view_signature)

    def test_loudness_controller_handle_drop_uses_shared_parser_and_skips_duplicates(self):
        first_path = os.path.join(self.tempdir.name, "first.wav")
        second_path = os.path.join(self.tempdir.name, "second.mp3")
        with open(first_path, "wb") as handle:
            handle.write(b"audio")
        with open(second_path, "wb") as handle:
            handle.write(b"audio")

        app = self.make_app_stub()
        app.state.loudness_files = [first_path]
        app.tasks.parse_dnd_files = lambda _data: [first_path, second_path, "/tmp/missing.flac"]
        controller = LoudnessController(app)
        controller._pop_l_tree = lambda: setattr(app, "loudness_tree_reloaded", True)
        controller._apply_workflow_state = lambda state: setattr(app, "workflow_state", state)
        controller.log = app.logs.append
        controller.toast = app.toasts.append

        controller.handle_drop(SimpleNamespace(data="ignored"))

        self.assertEqual([first_path, second_path], app.state.loudness_files)
        self.assertTrue(app.loudness_tree_reloaded)
        self.assertEqual("1 DATEIEN HINZUGEFÜGT", app.toasts[0])

    def test_release_title_matching_requires_exact_normalized_match(self):
        source_path = os.path.join(self.tempdir.name, "Intro.wav")
        with open(source_path, "wb") as handle:
            handle.write(b"audio")

        title_work = release_tools.find_work_by_title_like_audio_path(
            [
                {
                    "title": "Intro Remix",
                    "duration": "",
                    "production": "",
                    "year": "",
                    "audio_path": "",
                }
            ],
            source_path,
        )
        track = release_workflows.build_release_track_from_match(
            source_path,
            title_work=title_work,
        )

        self.assertIsNone(title_work)
        self.assertEqual("Datei", track["source"])
        self.assertEqual("Intro", track["title"])

    def test_release_title_matching_keeps_exact_normalized_matches(self):
        source_path = os.path.join(self.tempdir.name, "01 - Intro_matched.wav")
        with open(source_path, "wb") as handle:
            handle.write(b"audio")

        title_work = release_tools.find_work_by_title_like_audio_path(
            [
                {
                    "title": "Intro",
                    "duration": "1:11",
                    "production": "Prod",
                    "year": "2026",
                    "audio_path": "",
                }
            ],
            source_path,
        )
        track = release_workflows.build_release_track_from_match(
            source_path,
            title_work=title_work,
        )

        self.assertEqual("Datei→Werk", track["source"])
        self.assertEqual("Intro", track["title"])
        self.assertEqual("1:11", track["duration"])

    def test_root_mousewheel_routes_scroll_to_nearest_scrollable_panel(self):
        registry = {}
        panel = FakeScrollablePanelWidget("panel", registry)
        label = FakeWidgetNode("label", registry, parent=panel)
        app = SimpleNamespace(winfo_containing=lambda *_args: label)

        with mock.patch.object(
            akm_app.ui_patterns,
            "AkmScrollablePanel",
            FakeScrollablePanelWidget,
        ):
            akm_app.AKMApp._on_root_mousewheel(
                app,
                SimpleNamespace(x_root=12, y_root=18, num=4, delta=0),
            )

        self.assertEqual([(-1, "units")], panel.canvas.scroll_calls)

    def test_root_mousewheel_prefers_native_scrollable_widgets(self):
        registry = {}
        listbox = FakeNativeScrollWidget("listbox", registry)
        app = SimpleNamespace(winfo_containing=lambda *_args: listbox)

        with mock.patch.object(akm_app.tk, "Listbox", FakeNativeScrollWidget):
            akm_app.AKMApp._on_root_mousewheel(
                app,
                SimpleNamespace(x_root=8, y_root=10, num=0, delta=120),
            )

        self.assertEqual([(-1, "units")], listbox.scroll_calls)

    def test_merge_export_notes_replaces_previous_auto_lines(self):
        merged = release_tools.merge_export_notes(
            "Eigene Notiz\nMatched Export: alt.wav\nLimiter: Nein",
            "/tmp/neu.wav",
            True,
        )

        self.assertEqual(
            "Eigene Notiz\n\nMatched Export: /tmp/neu.wav\nLimiter: Ja",
            merged,
        )

    def test_start_distro_export_removes_old_release_artifacts_only(self):
        source_dir = os.path.join(self.tempdir.name, "source")
        export_root = os.path.join(self.tempdir.name, "exports")
        os.makedirs(source_dir, exist_ok=True)
        os.makedirs(export_root, exist_ok=True)

        cover_path = os.path.join(source_dir, "cover.jpg")
        audio_path = os.path.join(source_dir, "track.wav")
        with open(cover_path, "wb") as handle:
            handle.write(b"cover")
        with open(audio_path, "wb") as handle:
            handle.write(b"audio")

        release_dir = os.path.join(export_root, "My Release")
        os.makedirs(release_dir, exist_ok=True)
        obsolete_audio = os.path.join(release_dir, "09 - Old Track.wav")
        obsolete_cover = os.path.join(release_dir, "old_cover.jpg")
        obsolete_info = os.path.join(release_dir, "release_info.txt")
        obsolete_excel = os.path.join(release_dir, "tracklist.xlsx")
        keep_file = os.path.join(release_dir, "user_notes.txt")
        for path in (obsolete_audio, obsolete_cover, obsolete_info, keep_file):
            with open(path, "wb") as handle:
                handle.write(b"old")
        old_workbook = Workbook()
        old_workbook.save(obsolete_excel)
        old_workbook.close()

        ok, _status = release_workflows.start_distro_export(
            {
                "title": "My Release",
                "artist": "Artist",
                "export_dir": export_root,
                "cover_path": cover_path,
                "type": "Album",
                "release_date": "2026-04-05",
                "genre": "Pop",
                "subgenre": "Indie",
                "label": "Label",
                "copyright_line": "C 2026",
            },
            [
                {
                    "title": "Fresh Track",
                    "duration": "3:45",
                    "production": "Prod",
                    "year": "2026",
                    "audio_path": audio_path,
                }
            ],
        )

        self.assertTrue(ok)
        self.assertFalse(os.path.exists(obsolete_audio))
        self.assertFalse(os.path.exists(obsolete_cover))
        self.assertTrue(os.path.exists(obsolete_info))
        self.assertTrue(os.path.exists(keep_file))
        self.assertTrue(os.path.exists(os.path.join(release_dir, "01 - Fresh Track.wav")))
        self.assertTrue(os.path.exists(os.path.join(release_dir, "cover.jpg")))
        self.assertTrue(os.path.exists(os.path.join(release_dir, "tracklist.csv")))
        self.assertTrue(os.path.exists(os.path.join(release_dir, "tracklist.xlsx")))
        self.assertTrue(os.path.exists(os.path.join(release_dir, "checklist.txt")))
        with open(obsolete_info, "r", encoding="utf-8") as handle:
            self.assertIn("Release Title: My Release", handle.read())
        workbook = load_workbook(os.path.join(release_dir, "tracklist.xlsx"), read_only=True, data_only=True)
        try:
            sheet_names = workbook.sheetnames
            import_sheet = workbook["AKM Import"]
            info_sheet = workbook["Release Info"]
            track_sheet = workbook["Tracklist"]
            import_rows = list(import_sheet.iter_rows(values_only=True))
            info_rows = list(info_sheet.iter_rows(values_only=True))
            track_rows = list(track_sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        self.assertEqual("AKM Import", sheet_names[0])
        self.assertEqual(
            (
                "Title",
                "Duration",
                "Composer",
                "Production",
                "Year",
                "Notes",
                "Audio Path",
                "Track Number",
                "Release Title",
                "Artist",
                "Type",
                "Release Date",
                "Genre",
                "Subgenre",
                "Label",
                "Copyright",
                "Cover Path",
                "Source",
            ),
            import_rows[0],
        )
        self.assertEqual("Fresh Track", import_rows[1][0])
        self.assertEqual("My Release", import_rows[1][8])
        self.assertEqual(audio_path, import_rows[1][6])
        self.assertEqual(("Field", "Value"), info_rows[0])
        self.assertIn(("Release Title", "My Release"), info_rows)
        self.assertEqual(
            ("Track Number", "Title", "Duration", "Production", "Year", "Audio Path"),
            track_rows[0],
        )
        self.assertEqual("Fresh Track", track_rows[1][1])

    def test_loudness_tools_ensure_ffmpeg_in_path_uses_platform_separator(self):
        with mock.patch.object(loudness_tools, "_ffmpeg_search_locations", return_value=["C:\\ffmpeg\\bin"]), mock.patch.object(
            loudness_tools.os.path,
            "exists",
            side_effect=lambda value: value == "C:\\ffmpeg\\bin",
        ), mock.patch.dict(
            loudness_tools.os.environ,
            {"PATH": "C:\\Windows\\System32"},
            clear=False,
        ), mock.patch.object(loudness_tools.os, "pathsep", ";"):
            loudness_tools._ensure_ffmpeg_in_path()
            self.assertEqual("C:\\ffmpeg\\bin;C:\\Windows\\System32", loudness_tools.os.environ["PATH"])

    def test_open_in_finder_uses_windows_explorer_selection_for_files(self):
        with mock.patch.object(widgets.platform, "system", return_value="Windows"), mock.patch.object(
            widgets.os.path,
            "isdir",
            return_value=False,
        ), mock.patch.object(
            widgets.os.path,
            "normpath",
            return_value="C:\\tmp\\test.wav",
        ), mock.patch.object(widgets.subprocess, "run") as run_mock:
            ok, error = widgets.open_in_finder("C:/tmp/test.wav")

        self.assertTrue(ok)
        self.assertIsNone(error)
        run_mock.assert_called_once_with(["explorer", "/select,C:\\tmp\\test.wav"], check=False)


if __name__ == "__main__":
    unittest.main()
