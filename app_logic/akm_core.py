import csv
import json
import os
import tempfile
from datetime import datetime
from itertools import chain
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app_logic.config import cfg

# Use Centralized Config Constants
DATA_DIR = cfg.DATA_DIR
PROJECTS_DIR = cfg.PROJECTS_DIR
DATA_FILE = cfg.DATA_FILE
BACKUP_FILE = cfg.BACKUP_FILE
LANG_FILE = os.path.join(DATA_DIR, "lang.txt")
SETTINGS_FILE = cfg.SETTINGS_FILE
SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

STATUS_KEYS = ["in_progress", "ready", "submitted", "confirmed"]
DEFAULT_SETTINGS = {
    "release_default_artist": "Uwe Felchle Project",
}

LANGUAGES = {
    "de": {
        "status_map": {
            "in_progress": "🟡 in Arbeit",
            "ready": "🔵 bereit",
            "submitted": "🟢 gemeldet",
            "confirmed": "✅ bestätigt",
        },
        "status_keys": STATUS_KEYS,
    },
    "en": {
        "status_map": {
            "in_progress": "🟡 in progress",
            "ready": "🔵 ready",
            "submitted": "🟢 submitted",
            "confirmed": "✅ confirmed",
        },
        "status_keys": STATUS_KEYS,
    },
}


class DataFileError(RuntimeError):
    pass


def _today():
    return datetime.now().strftime("%Y-%m-%d")


def _ensure_storage_dir():
    cfg.ensure_dirs()


def _write_text_atomic(path, content):
    directory = os.path.dirname(path) or "."
    os.makedirs(directory, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(
        prefix=f".{os.path.basename(path)}.",
        suffix=".tmp",
        dir=directory,
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def _write_json_atomic(path, payload):
    directory = os.path.dirname(path) or "."
    os.makedirs(directory, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(
        prefix=f".{os.path.basename(path)}.",
        suffix=".tmp",
        dir=directory,
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, ensure_ascii=False)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def _clean_text(value):
    if value is None:
        return ""

    if value != value:
        return ""

    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _read_json_list(path, strict=False):
    if not os.path.exists(path):
        return []

    try:
        with open(path, "r", encoding="utf-8") as handle:
            raw_data = json.load(handle)
    except OSError as exc:
        if strict:
            raise DataFileError(
                f"{os.path.basename(path)} konnte nicht gelesen werden: {exc}"
            ) from exc
        return []
    except json.JSONDecodeError as exc:
        if strict:
            raise DataFileError(
                f"{os.path.basename(path)} ist beschädigt (Zeile {exc.lineno}, Spalte {exc.colno})."
            ) from exc
        return []

    if not isinstance(raw_data, list):
        if strict:
            raise DataFileError(f"{os.path.basename(path)} hat ein ungültiges Format.")
        return []

    return raw_data


def _read_json_dict(path, strict=False):
    if not os.path.exists(path):
        return {}

    try:
        with open(path, "r", encoding="utf-8") as handle:
            raw_data = json.load(handle)
    except OSError as exc:
        if strict:
            raise DataFileError(
                f"{os.path.basename(path)} konnte nicht gelesen werden: {exc}"
            ) from exc
        return {}
    except json.JSONDecodeError as exc:
        if strict:
            raise DataFileError(
                f"{os.path.basename(path)} ist beschädigt (Zeile {exc.lineno}, Spalte {exc.colno})."
            ) from exc
        return {}

    if not isinstance(raw_data, dict):
        if strict:
            raise DataFileError(f"{os.path.basename(path)} hat ein ungültiges Format.")
        return {}

    return raw_data


def _normalize_entry(entry):
    item = dict(entry) if isinstance(entry, dict) else {}
    item["title"] = _clean_text(item.get("title"))
    item["status"] = item.get("status", "in_progress")
    if item["status"] not in STATUS_KEYS:
        item["status"] = "in_progress"
    item["date"] = _clean_text(item.get("date")) or _today()
    item["last_change"] = _clean_text(item.get("last_change")) or item["date"]
    item["duration"] = _clean_text(item.get("duration"))
    item["composer"] = _clean_text(item.get("composer"))
    item["production"] = _clean_text(item.get("production"))
    item["year"] = _clean_text(item.get("year"))
    item["instrumental"] = bool(item.get("instrumental", False))
    item["notes"] = _clean_text(item.get("notes"))
    item["audio_path"] = _clean_text(item.get("audio_path"))
    tags = item.get("tags", [])
    if isinstance(tags, str):
        tags = [part.strip() for part in tags.split(",") if part.strip()]
    elif not isinstance(tags, list):
        tags = []
    item["tags"] = [_clean_text(tag) for tag in tags if _clean_text(tag)]
    return item


def _find_column(columns, candidates):
    normalized = {str(column).strip().lower(): column for column in columns}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    return None


def _get_row_value(values, index):
    if index is None or index >= len(values):
        return ""
    return values[index]


def load_excel_tracks(file_path):
    extension = os.path.splitext(file_path)[1].lower()
    if extension not in SUPPORTED_EXCEL_EXTENSIONS:
        raise DataFileError(
            "Nicht unterstütztes Excel-Format. Bitte Datei als .xlsx speichern."
        )

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise DataFileError(
            "Nicht unterstütztes Excel-Format. Bitte Datei als .xlsx speichern."
        ) from exc
    except BadZipFile as exc:
        raise DataFileError(
            "Excel-Datei ist beschädigt oder kein gültiges .xlsx-Format."
        ) from exc
    except OSError as exc:
        raise DataFileError(f"Excel-Datei konnte nicht gelesen werden: {exc}") from exc

    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row is None:
            return []

        first_values = list(first_row)
        headers = [_clean_text(value) for value in first_values]
        header_markers = {
            "titel",
            "title",
            "dauer",
            "duration",
            "length",
            "komponist",
            "composer",
        }
        has_header = any(header.lower() in header_markers for header in headers if header)

        if has_header:
            title_header = _find_column(headers, ["titel", "title"])
            duration_header = _find_column(headers, ["dauer", "duration", "length"])
            composer_header = _find_column(headers, ["komponist", "composer"])
            production_header = _find_column(headers, ["produktion", "production"])
            year_header = _find_column(headers, ["jahr", "year"])
            notes_header = _find_column(headers, ["notiz", "notizen", "notes"])
            audio_path_header = _find_column(
                headers,
                ["audio path", "audio-pfad", "audio_path", "audio", "pfad", "file path", "filepath"],
            )
            title_index = headers.index(title_header) if title_header is not None else 0
            duration_index = headers.index(duration_header) if duration_header is not None else None
            composer_index = headers.index(composer_header) if composer_header is not None else None
            production_index = headers.index(production_header) if production_header is not None else None
            year_index = headers.index(year_header) if year_header is not None else None
            notes_index = headers.index(notes_header) if notes_header is not None else None
            audio_path_index = headers.index(audio_path_header) if audio_path_header is not None else None
            data_rows = rows
        else:
            title_index = 0
            duration_index = 1 if len(first_values) > 1 else None
            composer_index = 2 if len(first_values) > 2 else None
            production_index = 3 if len(first_values) > 3 else None
            year_index = 4 if len(first_values) > 4 else None
            notes_index = 5 if len(first_values) > 5 else None
            audio_path_index = 6 if len(first_values) > 6 else None
            data_rows = chain((tuple(first_values),), rows)

        tracks = []
        for row in data_rows:
            values = list(row)
            title = _clean_text(_get_row_value(values, title_index))
            if not title:
                continue

            tracks.append(
                {
                    "title": title,
                    "duration": _clean_text(_get_row_value(values, duration_index)),
                    "composer": _clean_text(_get_row_value(values, composer_index)),
                    "production": _clean_text(_get_row_value(values, production_index)),
                    "year": _clean_text(_get_row_value(values, year_index)),
                    "notes": _clean_text(_get_row_value(values, notes_index)),
                    "audio_path": _clean_text(_get_row_value(values, audio_path_index)),
                }
            )
        return tracks
    finally:
        workbook.close()


def get_lang():
    if os.path.exists(LANG_FILE):
        with open(LANG_FILE, "r", encoding="utf-8") as handle:
            lang = handle.read().strip()
            if lang in LANGUAGES:
                return lang
    return "de"


def set_lang(lang):
    _ensure_storage_dir()
    if lang not in LANGUAGES:
        lang = "de"
    _write_text_atomic(LANG_FILE, lang)


def load_settings(strict=False):
    raw_settings = _read_json_dict(SETTINGS_FILE, strict=strict)
    settings = dict(DEFAULT_SETTINGS)

    if "release_default_artist" in raw_settings:
        settings["release_default_artist"] = _clean_text(
            raw_settings.get("release_default_artist")
        )

    return settings


def save_settings(settings):
    _ensure_storage_dir()
    merged = dict(DEFAULT_SETTINGS)
    if "release_default_artist" in settings:
        merged["release_default_artist"] = _clean_text(
            settings.get("release_default_artist")
        )
    _write_json_atomic(SETTINGS_FILE, merged)


def get_release_default_artist():
    settings = load_settings()
    return settings.get("release_default_artist", DEFAULT_SETTINGS["release_default_artist"])


def set_release_default_artist(value):
    settings = load_settings()
    settings["release_default_artist"] = _clean_text(value)
    save_settings(settings)


def load_data(strict=False):
    raw_data = _read_json_list(DATA_FILE, strict=strict)
    data = []
    for item in raw_data:
        normalized = _normalize_entry(item)
        if normalized["title"]:
            data.append(normalized)
    return data


def save_data(data):
    _ensure_storage_dir()
    normalized = []
    for item in data:
        entry = _normalize_entry(item)
        if entry["title"]:
            normalized.append(entry)
    _write_json_atomic(DATA_FILE, normalized)


def backup_data():
    _ensure_storage_dir()
    data = load_data(strict=True)
    _write_json_atomic(BACKUP_FILE, data)


def restore_data():
    if not os.path.exists(BACKUP_FILE):
        return False

    data = _read_json_list(BACKUP_FILE, strict=True)
    save_data(data)
    return True


def normalize(title):
    return _clean_text(title).lower()


def find_entry(data, title):
    norm = normalize(title)
    for item in data:
        if normalize(item["title"]) == norm:
            return item
    return None


def get_last_open():
    data = load_data(strict=True)
    open_entries = [item for item in data if item.get("status") != "confirmed"]
    if not open_entries:
        return None
    return max(
        open_entries,
        key=lambda item: (
            item.get("last_change", "") or item.get("date", ""),
            item.get("date", ""),
        ),
    )


def add_entry(
    title,
    duration="",
    composer="",
    production="",
    year="",
    instrumental=False,
    notes="",
    audio_path="",
    tags=None,
    status="in_progress",
):
    title = _clean_text(title)
    if not title:
        return False, "empty_title"

    status = _clean_text(status) or "in_progress"
    if status not in STATUS_KEYS:
        return False, "invalid_status"

    data = load_data(strict=True)
    existing = find_entry(data, title)
    if existing:
        return False, "already_exists"

    now = _today()
    clean_tags = []
    for tag in tags or []:
        cleaned = _clean_text(tag)
        if cleaned:
            clean_tags.append(cleaned)

    entry = {
        "title": title,
        "status": status,
        "date": now,
        "last_change": now,
        "duration": _clean_text(duration),
        "composer": _clean_text(composer),
        "production": _clean_text(production),
        "year": _clean_text(year),
        "instrumental": bool(instrumental),
        "notes": _clean_text(notes),
        "audio_path": _clean_text(audio_path),
        "tags": clean_tags,
    }
    data.append(entry)
    save_data(data)
    return True, entry


def update_status(title, status, lang):
    del lang
    return update_entry(title, {"status": status})


def import_tracks(tracks, status="ready"):
    if status not in STATUS_KEYS:
        status = "ready"

    data = load_data(strict=True)
    touched = []
    changed = False
    now = _today()

    for track in tracks or []:
        title = _clean_text(track.get("title"))
        if not title:
            continue
        duration = _clean_text(track.get("duration"))
        composer = _clean_text(track.get("composer"))
        production = _clean_text(track.get("production"))
        year = _clean_text(track.get("year"))
        notes = _clean_text(track.get("notes"))
        audio_path = _clean_text(track.get("audio_path"))

        existing = find_entry(data, title)
        if existing:
            entry_changed = False

            if duration and duration != existing.get("duration", ""):
                existing["duration"] = duration
                entry_changed = True
            if composer and composer != existing.get("composer", ""):
                existing["composer"] = composer
                entry_changed = True
            if production and production != existing.get("production", ""):
                existing["production"] = production
                entry_changed = True
            if year and year != existing.get("year", ""):
                existing["year"] = year
                entry_changed = True
            if notes and notes != existing.get("notes", ""):
                existing["notes"] = notes
                entry_changed = True
            if audio_path and audio_path != existing.get("audio_path", ""):
                existing["audio_path"] = audio_path
                entry_changed = True
            if existing["status"] == "in_progress" and status != "in_progress":
                existing["status"] = status
                entry_changed = True

            if entry_changed:
                existing["last_change"] = now
                changed = True
                action = "updated"
            else:
                action = "unchanged"

            touched.append(
                {
                    "title": title,
                    "duration": existing.get("duration", ""),
                    "composer": existing.get("composer", ""),
                    "production": existing.get("production", ""),
                    "year": existing.get("year", ""),
                    "notes": existing.get("notes", ""),
                    "audio_path": existing.get("audio_path", ""),
                    "status": existing["status"],
                    "action": action,
                }
            )
            continue

        entry = {
            "title": title,
            "status": status,
            "date": now,
            "last_change": now,
            "duration": duration,
            "composer": composer,
            "production": production,
            "year": year,
            "instrumental": False,
            "notes": notes,
            "audio_path": audio_path,
            "tags": [],
        }
        data.append(entry)
        changed = True
        touched.append(
            {
                "title": title,
                "duration": duration,
                "composer": composer,
                "production": production,
                "year": year,
                "notes": notes,
                "audio_path": audio_path,
                "status": status,
                "action": "added",
            }
        )

    if changed:
        save_data(data)

    return touched


def import_excel(file_path):
    tracks = load_excel_tracks(file_path)
    if not tracks:
        return []
    return import_tracks(tracks)


def get_status_map(lang):
    return LANGUAGES[lang]["status_map"]


def get_status_keys(lang):
    return LANGUAGES[lang]["status_keys"]


def get_all_entries():
    return load_data(strict=True)


def export_csv(lang, path=None):
    data = load_data(strict=True)
    status_map = get_status_map(lang)

    if not path:
        path = os.path.expanduser("~/Desktop/akm_export.csv")

    export_dir = os.path.dirname(path)
    if export_dir:
        os.makedirs(export_dir, exist_ok=True)

    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "title",
                "status",
                "date",
                "last_change",
                "duration",
                "composer",
                "production",
                "year",
                "instrumental",
                "notes",
                "audio_path",
                "tags",
            ],
        )
        writer.writeheader()
        for item in data:
            writer.writerow(
                {
                    "title": item["title"],
                    "status": status_map[item["status"]],
                    "date": item["date"],
                    "last_change": item.get("last_change", ""),
                    "duration": item.get("duration", ""),
                    "composer": item.get("composer", ""),
                    "production": item.get("production", ""),
                    "year": item.get("year", ""),
                    "instrumental": "yes" if item.get("instrumental") else "no",
                    "notes": item.get("notes", ""),
                    "audio_path": item.get("audio_path", ""),
                    "tags": ", ".join(item.get("tags", [])),
                }
            )
    return path


def update_entry(title, updates):
    data = load_data(strict=True)
    entry = find_entry(data, title)
    if not entry:
        return False, "not_found"

    allowed_fields = {
        "title",
        "duration",
        "composer",
        "production",
        "year",
        "instrumental",
        "notes",
        "audio_path",
        "tags",
        "status",
    }

    normalized_title = None
    if "title" in (updates or {}):
        normalized_title = _clean_text(updates.get("title"))
        if not normalized_title:
            return False, "empty_title"

        existing = find_entry(data, normalized_title)
        if existing is not None and existing is not entry:
            return False, "already_exists"

    changed = False
    for key, value in (updates or {}).items():
        if key not in allowed_fields:
            continue

        if key == "title":
            normalized_value = normalized_title
        elif key == "instrumental":
            normalized_value = bool(value)
        elif key == "tags":
            normalized_value = []
            for tag in value or []:
                cleaned = _clean_text(tag)
                if cleaned:
                    normalized_value.append(cleaned)
        elif key == "status":
            normalized_value = value if value in STATUS_KEYS else entry.get("status", "in_progress")
        else:
            normalized_value = _clean_text(value)

        if entry.get(key) != normalized_value:
            entry[key] = normalized_value
            changed = True

    if not changed:
        return True, entry

    entry["last_change"] = _today()
    save_data(data)
    return True, entry



def get_dashboard_stats():
    from app_logic import overview_tools

    return overview_tools.build_dashboard_stats(load_data(strict=True))

def save_project(path, data, cover_state=None, release_state=None, settings=None):
    """
    Saves the entire project state to a single JSON bundle.
    """
    bundle = {
        "version": "1.0",
        "timestamp": datetime.now().isoformat(),
        "data": data,
        "cover": cover_state or {},
        "release": release_state or {},
        "settings": settings or {}
    }
    _write_json_atomic(path, bundle)
    return True

def load_project(path):
    """
    Loads a project bundle from a JSON file.
    """
    if not os.path.exists(path):
        return None

    name = os.path.basename(path)
    try:
        with open(path, "r", encoding="utf-8") as handle:
            bundle = json.load(handle)
    except OSError as exc:
        raise DataFileError(f"{name} konnte nicht gelesen werden: {exc}") from exc
    except json.JSONDecodeError as exc:
        raise DataFileError(
            f"{name} ist beschädigt (Zeile {exc.lineno}, Spalte {exc.colno})."
        ) from exc

    if not isinstance(bundle, dict):
        raise DataFileError(f"{name} hat ein ungültiges Projektformat.")

    if "data" in bundle and not isinstance(bundle.get("data"), list):
        raise DataFileError(f"{name} enthält ein ungültiges Datenformat.")
    for key in ("cover", "release", "settings"):
        if key in bundle and not isinstance(bundle.get(key), dict):
            raise DataFileError(f"{name} enthält einen ungültigen Bereich: {key}.")

    return bundle
