import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app_logic import release_tools
from app_logic.text_utils import clean_mapping_values, clean_text as _clean_text


RELEASE_TRACK_FIELDS = (
    "title",
    "duration",
    "composer",
    "production",
    "year",
    "notes",
    "audio_path",
    "source",
)
RELEASE_IMPORT_FIELDS = RELEASE_TRACK_FIELDS[:-1]
RELEASE_INFO_FIELDS = (
    ("title", "Release Title"),
    ("artist", "Artist"),
    ("type", "Type"),
    ("release_date", "Release Date"),
    ("genre", "Genre"),
    ("subgenre", "Subgenre"),
    ("label", "Label"),
    ("copyright_line", "Copyright"),
    ("cover_path", "Cover Path"),
)
RELEASE_INFO_KEYS = tuple(key for key, _label in RELEASE_INFO_FIELDS)


def _build_release_track(base_track=None, **overrides):
    track = clean_mapping_values(base_track, RELEASE_TRACK_FIELDS)
    for key, value in overrides.items():
        track[key] = _clean_text(value)
    return track


def clean_release_drop_paths(raw_paths):
    paths = []
    seen = set()
    for raw in raw_paths or []:
        cleaned = _clean_text(raw).strip("{}")
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        paths.append(cleaned)
    return paths


def copy_work_to_release_track(item, audio_path=None, source="Werk"):
    track = _build_release_track(item, source=source)
    if audio_path is not None:
        track["audio_path"] = _clean_text(audio_path)
    return track


def build_file_release_track(path, duration=""):
    title = os.path.splitext(os.path.basename(path or ""))[0]
    return _build_release_track(
        title=title,
        duration=duration,
        audio_path=path,
        source="Datei",
    )


def build_release_track_from_match(path, exact_work=None, title_work=None, duration=""):
    if exact_work is not None:
        return copy_work_to_release_track(
            exact_work,
            audio_path=path,
            source="Werk",
        )
    if title_work is not None:
        return copy_work_to_release_track(
            title_work,
            audio_path=path,
            source="Datei→Werk",
        )
    return build_file_release_track(path, duration=duration)


def collect_release_track_identities(tracks):
    identities = set()
    for item in tracks or []:
        identity = release_tools.release_track_identity(item)
        if identity is not None:
            identities.add(identity)
    return identities


def append_unique_release_tracks(existing_tracks, candidate_tracks):
    updated_tracks = list(existing_tracks or [])
    existing_identities = collect_release_track_identities(updated_tracks)
    added_tracks = []
    duplicate_tracks = []

    for track in candidate_tracks or []:
        identity = release_tools.release_track_identity(track)
        if identity is not None and identity in existing_identities:
            duplicate_tracks.append(track)
            continue
        updated_tracks.append(track)
        added_tracks.append(track)
        if identity is not None:
            existing_identities.add(identity)

    return {
        "tracks": updated_tracks,
        "added": added_tracks,
        "duplicates": duplicate_tracks,
    }


def remove_release_track_at(tracks, index):
    if index is None or index < 0 or index >= len(tracks or []):
        return list(tracks or []), None

    updated_tracks = list(tracks or [])
    removed = updated_tracks.pop(index)
    return updated_tracks, removed


def move_release_track(tracks, index, direction):
    updated_tracks = list(tracks or [])
    if index is None or index < 0 or index >= len(updated_tracks):
        return updated_tracks, None

    target_index = index + direction
    if target_index < 0 or target_index >= len(updated_tracks):
        return updated_tracks, None

    updated_tracks[index], updated_tracks[target_index] = (
        updated_tracks[target_index],
        updated_tracks[index],
    )
    return updated_tracks, target_index


def safe_release_directory_name(title):
    safe_name = "".join(
        ch if ch.isalnum() or ch in "-_ " else "_"
        for ch in _clean_text(title)
    ).strip()
    return safe_name or "Release"


def build_release_export_preserve_paths(cover_path, tracks):
    preserve_paths = []
    if _clean_text(cover_path):
        preserve_paths.append(_clean_text(cover_path))
    preserve_paths.extend(
        _clean_text(item.get("audio_path"))
        for item in (tracks or [])
        if _clean_text(item.get("audio_path"))
    )
    return preserve_paths


def build_release_info_lines(metadata):
    return [f"{label}: {value}" for label, value in build_release_info_rows(metadata)]


def build_release_info_rows(metadata):
    values = clean_mapping_values(metadata, RELEASE_INFO_KEYS)
    return [(label, values[key]) for key, label in RELEASE_INFO_FIELDS]


def build_release_track_csv_rows(tracks):
    rows = [["Track Number", "Title", "Duration", "Production", "Year", "Audio Path"]]
    for index, item in enumerate(tracks or [], start=1):
        rows.append(
            [
                index,
                _clean_text(item.get("title")),
                _clean_text(item.get("duration")),
                _clean_text(item.get("production")),
                _clean_text(item.get("year")),
                _clean_text(item.get("audio_path")),
            ]
        )
    return rows


def build_release_import_rows(metadata, tracks):
    values = clean_mapping_values(metadata, RELEASE_INFO_KEYS)
    rows = [[
        "Title",
        "Duration",
        "Composer",
        "Production",
        "Year",
        "Notes",
        "Audio Path",
        "Track Number",
        "Release Title",
        "Artist",
        "Type",
        "Release Date",
        "Genre",
        "Subgenre",
        "Label",
        "Copyright",
        "Cover Path",
        "Source",
    ]]
    for index, item in enumerate(tracks or [], start=1):
        rows.append(
            [
                _clean_text(item.get("title")),
                _clean_text(item.get("duration")),
                _clean_text(item.get("composer")),
                _clean_text(item.get("production")),
                _clean_text(item.get("year")),
                _clean_text(item.get("notes")),
                _clean_text(item.get("audio_path")),
                index,
                values["title"],
                values["artist"],
                values["type"],
                values["release_date"],
                values["genre"],
                values["subgenre"],
                values["label"],
                values["copyright_line"],
                values["cover_path"],
                _clean_text(item.get("source")),
            ]
        )
    return rows


def build_release_import_tracks(tracks):
    importable_tracks = []
    for item in tracks or []:
        title = _clean_text((item or {}).get("title"))
        if not title:
            continue
        importable_tracks.append(clean_mapping_values(item, RELEASE_IMPORT_FIELDS))
    return importable_tracks


def _autosize_excel_columns(sheet, min_width=12, max_width=48):
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            value = _clean_text(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = max(
            min_width,
            min(max_length + 2, max_width),
        )


def write_release_track_workbook(output_path, metadata, tracks):
    workbook = Workbook()
    import_sheet = workbook.active
    import_sheet.title = "AKM Import"
    for row in build_release_import_rows(metadata, tracks):
        import_sheet.append(row)
    import_sheet.freeze_panes = "A2"

    info_sheet = workbook.create_sheet("Release Info")
    info_sheet.append(["Field", "Value"])
    for row in build_release_info_rows(metadata):
        info_sheet.append(list(row))
    info_sheet.freeze_panes = "A2"

    track_sheet = workbook.create_sheet("Tracklist")
    for row in build_release_track_csv_rows(tracks):
        track_sheet.append(row)
    track_sheet.freeze_panes = "A2"

    header_font = Font(bold=True)
    for sheet in (import_sheet, info_sheet, track_sheet):
        for cell in sheet[1]:
            cell.font = header_font
        _autosize_excel_columns(sheet)

    workbook.save(output_path)
    workbook.close()


def build_release_audio_target_name(index, item, audio_path):
    stem = os.path.splitext(os.path.basename(audio_path or ""))[0]
    ext = os.path.splitext(audio_path or "")[1]
    safe_track_name = "".join(
        ch if ch.isalnum() or ch in "-_ " else "_"
        for ch in _clean_text((item or {}).get("title")) or stem
    ).strip() or stem
    return f"{index:02d} - {safe_track_name}{ext}"


def build_release_checklist_lines(cover_path, tracks, copied_audio, missing_audio):
    return [
        f"Cover gesetzt: {'Ja' if _clean_text(cover_path) else 'Nein'}",
        f"Tracks im Release: {len(tracks or [])}",
        f"Audio-Dateien kopiert: {copied_audio}",
        f"Tracks ohne Audio-Pfad / Kopierfehler: {missing_audio}",
        "Lautheit schon geprüft: bitte prüfen",
        "Reihenfolge geprüft: bitte prüfen",
        "Release-Datum gesetzt: bitte prüfen",
    ]


def build_release_export_status_text(release_dir, tracks, copied_audio):
    return (
        f"Distro-Export bereit: {release_dir} | "
        f"Tracks: {len(tracks or [])} | "
        f"Audio kopiert: {copied_audio}"
    )

def start_distro_export(metadata, tracks):
    """
    Orchestrates the full distribution export: directory creation, 
    audio copying, and metadata generation.
    """
    import shutil
    import csv
    
    export_dir = metadata.get("export_dir", "").strip()
    if not export_dir:
        return False, "Kein Export-Ordner angegeben."
    
    release_title = metadata.get("title", "Release").strip()
    safe_title = safe_release_directory_name(release_title)
    full_export_path = os.path.join(export_dir, safe_title)
    
    try:
        os.makedirs(full_export_path, exist_ok=True)
        _, cleanup_errors = release_tools.cleanup_release_export_dir(full_export_path)
        if cleanup_errors:
            first_path, first_error = cleanup_errors[0]
            return False, f"Fehler beim Bereinigen von {first_path}: {first_error}"
        
        # 1. Copy Cover
        cover_src = metadata.get("cover_path", "").strip()
        if cover_src and os.path.exists(cover_src):
            ext = os.path.splitext(cover_src)[1]
            shutil.copy2(cover_src, os.path.join(full_export_path, f"cover{ext}"))

        # 2. Copy Tracks
        copied_audio = 0
        missing_audio = 0
        for i, track in enumerate(tracks, start=1):
            src = track.get("path") or track.get("audio_path")
            if src and os.path.exists(src):
                target_name = build_release_audio_target_name(i, track, src)
                shutil.copy2(src, os.path.join(full_export_path, target_name))
                copied_audio += 1
            else:
                missing_audio += 1

        # 3. Write Release Info
        info_lines = build_release_info_lines(metadata)
        with open(os.path.join(full_export_path, "release_info.txt"), "w", encoding="utf-8") as f:
            f.write("\n".join(info_lines))

        # 4. Write Tracklist CSV
        csv_rows = build_release_track_csv_rows(tracks)
        with open(os.path.join(full_export_path, "tracklist.csv"), "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(csv_rows)

        # 4b. Write Tracklist Excel workbook
        write_release_track_workbook(
            os.path.join(full_export_path, "tracklist.xlsx"),
            metadata,
            tracks,
        )

        # 5. Write Checklist
        checklist_lines = build_release_checklist_lines(cover_src, tracks, copied_audio, missing_audio)
        with open(os.path.join(full_export_path, "checklist.txt"), "w", encoding="utf-8") as f:
            f.write("\n".join(checklist_lines))

        status = build_release_export_status_text(full_export_path, tracks, copied_audio)
        return True, status

    except Exception as e:
        return False, f"Fehler beim Export: {str(e)}"
